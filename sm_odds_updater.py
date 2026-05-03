"""
Fetch SportsMarket matched odds and populate column V (SM Odds) in the master sheet.
Runs at 7am and 7pm via launchd, alongside the results updater.
"""

import logging
import re
import shutil
import zipfile
from pathlib import Path
from datetime import datetime

import openpyxl

from sportsmarket_api import fetch_all_orders, parse_order, match_orders_to_sheet

from constants import PROJECT_ROOT, MASTER_PATH, LOG_DIR
from master_io import master_lock, safe_save_workbook, check_free_space, InsufficientSpaceError
LOG_DIR.mkdir(parents=True, exist_ok=True)
LOG_PATH = LOG_DIR / f"sm_odds_{datetime.now().strftime('%Y-%m-%d')}.log"
SM_COL = 22  # column V

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(str(LOG_PATH)), logging.StreamHandler()],
    force=True,
)
logger = logging.getLogger("sm_odds_updater")


def main():
    logger.info("Starting SM odds updater...")

    if not MASTER_PATH.exists():
        logger.error(f"Master spreadsheet not found: {MASTER_PATH}")
        return

    # Fetch SM orders
    orders = fetch_all_orders()
    if not orders:
        logger.warning("No SM orders fetched — check session/credentials")
        return

    parsed = [p for p in (parse_order(o) for o in orders) if p is not None]
    logger.info(f"Parsed {len(parsed)} SM orders")

    # Load + write under the master lock to serialise with other writers.
    with master_lock(MASTER_PATH):
        try:
            wb = openpyxl.load_workbook(MASTER_PATH)
        except PermissionError:
            logger.warning("Master spreadsheet is locked (open in Excel?) — skipping SM odds update")
            return

        ws = wb['Master Bet Tracker']

        last = 1
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value is None:
                break
            last = r

        # Ensure header
        ws.cell(row=1, column=SM_COL, value="SM Odds")

        # Match and populate
        matches, unmatched = match_orders_to_sheet(parsed, ws, last)

        filled = 0
        for row, avg_odds, _ in matches:
            ws.cell(row=row, column=SM_COL, value=avg_odds)
            ws.cell(row=row, column=SM_COL).number_format = '0.000'
            filled += 1

        logger.info(f"Matched: {len(matches)}, Filled: {filled}, Unmatched: {len(unmatched)}")

        # Save
        try:
            safe_save_workbook(wb, MASTER_PATH)
            wb.close()
            logger.info(f"Saved: {MASTER_PATH}")
        except PermissionError:
            wb.close()
            logger.warning("Could not save — file locked by Excel")
            return
        except InsufficientSpaceError as e:
            wb.close()
            logger.error(f"Disk too full to save master safely ({e}) — skipping SM odds update; live file untouched")
            return

        # Fix sheet view (already atomic via tmp+move, but free-space pre-flight first)
        try:
            check_free_space(MASTER_PATH)
        except InsufficientSpaceError as e:
            logger.warning(f"Skipping sheet-view fix: {e}")
            logger.info("Done")
            return

        tmp = str(MASTER_PATH) + '.viewfix.tmp'
        try:
            with zipfile.ZipFile(str(MASTER_PATH), 'r') as zin, \
                 zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename == 'xl/worksheets/sheet1.xml':
                        text = data.decode('utf-8')
                        clean = ('<sheetViews><sheetView zoomScaleNormal="100" workbookViewId="0">'
                                 '<pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen" />'
                                 '<selection pane="bottomLeft" activeCell="A2" sqref="A2" />'
                                 '</sheetView></sheetViews>')
                        text = re.sub(r'<sheetViews>.*?</sheetViews>', clean, text, flags=re.DOTALL)
                        data = text.encode('utf-8')
                    zout.writestr(item, data)
            import os as _os
            _os.replace(tmp, str(MASTER_PATH))
        except Exception as e:
            logger.warning(f"Sheet view fix failed: {e}")
            try:
                from pathlib import Path as _P
                _P(tmp).unlink(missing_ok=True)
            except Exception:
                pass

    logger.info("Done")


if __name__ == "__main__":
    main()
