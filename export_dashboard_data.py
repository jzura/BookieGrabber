"""
Export master spreadsheet data to CSV for the Streamlit Cloud dashboard.
Runs every 12 hours via launchd, then pushes to GitHub.
"""

import logging
import subprocess
from pathlib import Path
from datetime import datetime

import openpyxl
import pandas as pd

from constants import PROJECT_ROOT, MASTER_PATH, DASHBOARD_DIR
CSV_PATH = DASHBOARD_DIR / "bets.csv"

LOG_DIR = PROJECT_ROOT / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)
LOG_PATH = LOG_DIR / f"dashboard_export_{datetime.now().strftime('%Y-%m-%d')}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(str(LOG_PATH)), logging.StreamHandler()],
    force=True,
)
logger = logging.getLogger("dashboard_export")


from strategy_config import compute_rpd as _rpd


def _compute_stake_and_return(df):
    """Compute Stake, Return, Profit from raw data (since Excel formulas aren't cached)."""
    from collections import defaultdict

    df = df.copy()
    df['Stake'] = None
    df['Return'] = None
    df['Profit'] = None

    from strategy_config import (is_core_qualifying as _is_core_q,
                                  is_btts_fade as _is_btts_f,
                                  is_15g_fade as _is_15g_f)

    def is_core(row):
        try:
            bf = float(row['BF']); vol = float(row['Volume'])
        except (ValueError, TypeError):
            return False
        rpd = _rpd(row['Bet365'], bf)
        return _is_core_q(row['Market'], row['Prediction'], bf, vol, rpd)

    def is_btts_fade(row):
        try:
            vol = float(row['Volume'])
        except (ValueError, TypeError):
            return False
        rpd = _rpd(row['Bet365'], row['BF'])
        return _is_btts_f(row['Market'], row['Prediction'], rpd, vol)

    def is_15g_fade(row):
        try:
            vol = float(row['Volume'])
        except (ValueError, TypeError):
            return False
        rpd = _rpd(row['Bet365'], row['BF'])
        return _is_15g_f(row['Market'], row['Prediction'], rpd, vol)

    # Build core set and double-stake
    core_idx = set()
    match_count = defaultdict(int)
    for idx, row in df.iterrows():
        if is_core(row):
            core_idx.add(idx)
            mk = (str(row['Date']), str(row['Home']), str(row['Away']))
            match_count[mk] += 1

    # Double-stake: highest BF per match with RPD=1.0 and ≥2 core bets
    match_dbl = defaultdict(list)
    for idx in core_idx:
        row = df.loc[idx]
        rpd = _rpd(row['Bet365'], row['BF'])
        mk = (str(row['Date']), str(row['Home']), str(row['Away']))
        from strategy_config import DOUBLE_STAKE_RPD, DOUBLE_STAKE_MIN_COUNT
        if rpd is not None and rpd <= DOUBLE_STAKE_RPD and match_count[mk] >= DOUBLE_STAKE_MIN_COUNT:
            try:
                match_dbl[mk].append((idx, float(row['BF'])))
            except Exception:
                pass
    dbl_idx = set()
    for mk, candidates in match_dbl.items():
        best = max(candidates, key=lambda x: x[1])[0]
        dbl_idx.add(best)

    # Under 2.5G piggyback: matches where 1.5G Under qualifies as core
    matches_with_core_15g = set()
    for idx in core_idx:
        row = df.loc[idx]
        if row['Market'] == '1.5G' and row['Prediction'] == 0:
            matches_with_core_15g.add((str(row['Date']), str(row['Home']), str(row['Away'])))

    piggyback_idx = set()
    for idx, row in df.iterrows():
        from strategy_config import is_25g_piggyback as _is_pb
        mk = (str(row['Date']), str(row['Home']), str(row['Away']))
        try:
            bf = float(row['BF'])
        except (ValueError, TypeError):
            bf = 0
        if _is_pb(row['Market'], row['Prediction'], bf, mk in matches_with_core_15g):
            piggyback_idx.add(idx)

    # Assign stakes
    for idx, row in df.iterrows():
        stake = 0
        fade = False
        if idx in core_idx:
            stake = 2 if idx in dbl_idx else 1
        elif idx in piggyback_idx:
            stake = 1
        elif is_btts_fade(row):
            stake = 1; fade = True
        elif is_15g_fade(row):
            stake = 1; fade = True

        if stake == 0: continue

        result = row.get('Result')
        if result is None or pd.isna(result):
            df.at[idx, 'Stake'] = stake
            continue

        result = int(result)
        try:
            bf = float(row['BF'])
        except Exception:
            continue

        # Use SM matched odds when available, else estimate from BF with tiered discount
        from strategy_config import FADE_ODDS_HAIRCUT, estimate_sm_odds
        sm_odds = row.get('SM_Odds')
        try:
            sm_odds = float(sm_odds) if pd.notna(sm_odds) and sm_odds else None
        except (ValueError, TypeError):
            sm_odds = None

        if fade:
            if result == 0:
                if sm_odds:
                    # SM_Odds already includes commission — use directly
                    ret = stake * sm_odds
                else:
                    opp = 1 / (1 - 1 / bf) * (1 - FADE_ODDS_HAIRCUT)
                    c = 0.01 if opp <= 1.5 else 0.02 if opp <= 2.8 else 0.03 if opp <= 3.5 else 0.04
                    ret = stake * (1 + (opp - 1) * (1 - c))
            else:
                ret = 0
        else:
            if result == 1:
                if sm_odds:
                    # SM_Odds already includes commission — use directly
                    ret = stake * sm_odds
                else:
                    odds = estimate_sm_odds(bf)
                    c = 0.01 if odds <= 1.5 else 0.02 if odds <= 2.8 else 0.03 if odds <= 3.5 else 0.04
                    ret = stake * (1 + (odds - 1) * (1 - c))
            else:
                ret = 0

        df.at[idx, 'Stake'] = stake
        df.at[idx, 'Return'] = round(ret, 4)
        df.at[idx, 'Profit'] = round(ret - stake, 4)

    return df


def export_csv():
    if not MASTER_PATH.exists():
        logger.error(f"Master not found: {MASTER_PATH}")
        return False

    logger.info("Loading master spreadsheet...")
    try:
        wb = openpyxl.load_workbook(MASTER_PATH, data_only=True)
    except PermissionError:
        logger.warning("File locked — skipping export")
        return False

    ws = wb['Master Bet Tracker']

    rows = []
    for r in range(2, ws.max_row + 1):
        bt = ws.cell(row=r, column=1).value
        if bt is None:
            break
        d = ws.cell(row=r, column=2).value
        if isinstance(d, datetime):
            d = d.date()

        # Read raw values — formulas won't be cached
        result = ws.cell(row=r, column=16).value
        hg = ws.cell(row=r, column=14).value
        ag = ws.cell(row=r, column=15).value
        pred = ws.cell(row=r, column=8).value
        goals = ws.cell(row=r, column=13).value

        # Compute result if it's a formula (None in data_only mode)
        if result is None and hg is not None and ag is not None:
            try:
                hg_i, ag_i = int(float(hg)), int(float(ag))
                tg = hg_i + ag_i
                if bt == 'BTTS':
                    both = (hg_i > 0 and ag_i > 0)
                    result = (1 if both else 0) if pred == 1 else (0 if both else 1)
                elif bt in ('1.5G', '2.5G', '3.5G'):
                    line = float(bt.replace('G', ''))
                    result = (1 if tg > line else 0) if pred == 1 else (1 if tg < line else 0)
            except (ValueError, TypeError):
                pass
        elif result is None and goals is not None:
            try:
                tg = int(float(goals))
                if bt in ('1.5G', '2.5G', '3.5G'):
                    line = float(bt.replace('G', ''))
                    result = (1 if tg > line else 0) if pred == 1 else (1 if tg < line else 0)
            except (ValueError, TypeError):
                pass

        # Compute RPD
        o365 = ws.cell(row=r, column=9).value
        bf = ws.cell(row=r, column=10).value
        rpd = _rpd(o365, bf)

        rows.append({
            'Market': bt,
            'Date': d,
            'Home': ws.cell(row=r, column=3).value,
            'Away': ws.cell(row=r, column=4).value,
            'Competition': ws.cell(row=r, column=5).value,
            'Prediction': pred,
            'Bet365': o365,
            'BF': bf,
            'Volume': ws.cell(row=r, column=11).value,
            'RPD': rpd,
            'Goals': goals if isinstance(goals, (int, float)) else (int(hg) + int(ag) if hg is not None and ag is not None else None),
            'HG': hg,
            'AG': ag,
            'Result': result,
            'Stake': None,  # computed below
            'Return': None,
            'Profit': None,
            'SM_Odds': ws.cell(row=r, column=22).value,
        })
    wb.close()

    df = pd.DataFrame(rows)
    df = _compute_stake_and_return(df)

    DASHBOARD_DIR.mkdir(parents=True, exist_ok=True)
    df.to_csv(CSV_PATH, index=False)

    n_staked = df['Stake'].notna().sum()
    n_settled = df['Profit'].notna().sum()
    logger.info(f"Exported {len(df)} rows ({n_staked} staked, {n_settled} settled) to {CSV_PATH}")
    return True


def git_push():
    try:
        subprocess.run(["git", "add", str(DASHBOARD_DIR)], cwd=str(PROJECT_ROOT),
                       capture_output=True, timeout=30)
        result = subprocess.run(
            ["git", "commit", "-m", f"Dashboard data update {datetime.now().strftime('%Y-%m-%d %H:%M')}"],
            cwd=str(PROJECT_ROOT), capture_output=True, text=True, timeout=30)
        if "nothing to commit" in result.stdout:
            logger.info("No changes to push")
            return
        subprocess.run(["git", "push"], cwd=str(PROJECT_ROOT),
                       capture_output=True, timeout=60)
        logger.info("Pushed to GitHub")
    except Exception as e:
        logger.error(f"Git push failed: {e}")


def export_fx_rate():
    """Fetch and save EUR/AUD rate."""
    try:
        import json, requests as req
        r = req.get('https://api.exchangerate-api.com/v4/latest/EUR', timeout=10)
        if r.status_code == 200:
            rate = r.json()['rates']['AUD']
            fx_file = DASHBOARD_DIR / "fx_rate.json"
            fx_file.write_text(json.dumps({
                "EUR_AUD": rate,
                "updated": datetime.now().isoformat()
            }))
            logger.info(f"EUR/AUD rate: {rate}")
    except Exception as e:
        logger.warning(f"FX rate fetch failed: {e}")


def export_sm_balance():
    """Fetch and save SM account balance info."""
    try:
        import requests
        from sportsmarket_api import get_session, SM_BASE, SM_USERNAME
        token = get_session()
        if not token:
            logger.warning("No SM session — skipping balance export")
            return
        r = requests.get(f"{SM_BASE}/customers/{SM_USERNAME}/accounting_info/",
                        headers={"Accept": "application/json", "session": token,
                                 "x-molly-client-name": "sonic"}, timeout=15)
        if r.status_code == 200:
            data = r.json()
            balance_file = DASHBOARD_DIR / "sm_balance.json"
            import json
            # Append timestamped entry
            history = []
            if balance_file.exists():
                try:
                    history = json.loads(balance_file.read_text())
                except Exception:
                    pass
            entry = {"timestamp": datetime.now().isoformat()}
            for item in data.get("data", []):
                entry[item["key"]] = item["value"]
            history.append(entry)
            # Keep last 365 entries
            history = history[-365:]
            balance_file.write_text(json.dumps(history, indent=2))
            logger.info(f"SM balance: {entry.get('current_balance', '?')}")
    except Exception as e:
        logger.warning(f"SM balance export failed: {e}")


def export_odds_timeline_summary():
    """Consolidate all raw odds_timeline CSVs into a single summary CSV for the dashboard."""
    import glob
    timeline_dir = PROJECT_ROOT / "data" / "odds_timeline"
    summary_path = DASHBOARD_DIR / "odds_timeline_summary.csv"

    if not timeline_dir.exists():
        return

    files = glob.glob(str(timeline_dir / "*" / "*.csv"))
    if not files:
        return

    dfs = []
    for f in files:
        try:
            df = pd.read_csv(f, on_bad_lines="skip")
            # Add league from directory name
            league = Path(f).parent.name
            df["league"] = league
            dfs.append(df)
        except Exception:
            continue

    if not dfs:
        return

    combined = pd.concat(dfs, ignore_index=True)
    combined.to_csv(summary_path, index=False)
    logger.info(f"Odds timeline summary: {len(combined)} snapshots from {len(files)} files")


def backup_master_spreadsheet():
    """Save a daily backup of the master spreadsheet."""
    import shutil
    backup_dir = PROJECT_ROOT / "data" / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")
    backup_path = backup_dir / f"master_backup_{date_str}.xlsx"
    if not backup_path.exists() and MASTER_PATH.exists():
        try:
            shutil.copy2(MASTER_PATH, backup_path)
            logger.info(f"Backup saved: {backup_path}")
            # Keep only last 14 days of backups
            import glob
            backups = sorted(glob.glob(str(backup_dir / "master_backup_*.xlsx")))
            for old in backups[:-14]:
                Path(old).unlink()
        except Exception as e:
            logger.warning(f"Backup failed: {e}")


def main():
    logger.info("Starting dashboard data export...")
    backup_master_spreadsheet()
    if export_csv():
        export_fx_rate()
        export_sm_balance()
        export_odds_timeline_summary()
        git_push()
    logger.info("Done")


if __name__ == "__main__":
    main()