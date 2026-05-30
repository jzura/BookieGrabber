"""Audit the last 200 unique settled matches in the master tracker:
- Re-fetch each match from the Odds API
- Re-derive HG/AG using cross-check vs per-half sum
- Flag any row where stored HG/AG disagrees with the re-derived truth
"""
import os, sys, time
from pathlib import Path
from datetime import datetime, timedelta
from collections import defaultdict

sys.path.insert(0, "/Users/Joel/REPOS/BookieGrabber")
import requests
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv("/Users/Joel/REPOS/BookieGrabber/.env")
from results_updater import ODDS_API_LEAGUES

API_KEY = os.getenv("ODDS_API_KEY")
MASTER = "/Users/Joel/Desktop/EFB_Master_Bet_Tracker_VS Code.xlsx"


def derive(scores):
    """Same cross-check logic as the patched results_updater."""
    periods = scores.get("periods", {}) or {}
    ft = periods.get("fulltime", {}) or {}
    p1 = periods.get("p1", {}) or {}
    p2 = periods.get("p2", {}) or {}
    top_h, top_a = scores.get("home"), scores.get("away")
    ft_h, ft_a = ft.get("home"), ft.get("away")
    if p1 and p2 and "home" in p1 and "home" in p2:
        try:
            half_h = int(p1.get("home", 0)) + int(p2.get("home", 0))
            half_a = int(p1.get("away", 0)) + int(p2.get("away", 0))
            if ft_h == half_h and ft_a == half_a:
                return ft_h, ft_a, "ft_matches_halves"
            if top_h == half_h and top_a == half_a:
                return top_h, top_a, "top_matches_halves"
            return (top_h if top_h is not None else ft_h,
                    top_a if top_a is not None else ft_a,
                    "neither_matches_halves")
        except (TypeError, ValueError):
            pass
    return (top_h if top_h is not None else ft_h,
            top_a if top_a is not None else ft_a,
            "no_halves")


# 1. Read last 200 unique matches with results from master
wb = load_workbook(MASTER, read_only=True, data_only=True)
ws = wb.active
matches = {}  # (date, home, away, comp) -> {"hg":..,"ag":..,"row_idxs":[...]}
for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if row[0] is None: continue
    bt, dt, h, a, comp, _, _, _, _, _, _, _, _, hg, ag = row[:15]
    if dt is None or h is None or a is None: continue
    if hg is None or ag is None: continue
    try:
        d = dt.date() if hasattr(dt, "date") else None
    except Exception:
        d = None
    if not d: continue
    key = (d, str(h), str(a), str(comp))
    if key not in matches:
        matches[key] = {"hg": int(hg), "ag": int(ag), "rows": []}
    matches[key]["rows"].append(r_idx)
wb.close()

# Sort by date desc, take last 200
keys_sorted = sorted(matches.keys(), key=lambda k: k[0], reverse=True)[:200]
print(f"Auditing {len(keys_sorted)} most recent settled matches")
print(f"Date range: {keys_sorted[-1][0]} to {keys_sorted[0][0]}\n")

# 2. Group by (sport_key, month) for batched API queries
buckets = defaultdict(list)  # (sport_key, year, month) -> [keys]
skipped_unmapped = []
for key in keys_sorted:
    _, _, _, comp = key
    sk = ODDS_API_LEAGUES.get(comp)
    if not sk:
        skipped_unmapped.append(comp)
        continue
    d = key[0]
    buckets[(sk, d.year, d.month)].append(key)

print(f"Buckets: {len(buckets)} (sport_key, month) queries")
if skipped_unmapped:
    from collections import Counter
    cc = Counter(skipped_unmapped)
    print(f"Skipped {len(skipped_unmapped)} matches with unmapped competition: {dict(cc)}")
print()

# 3. Fetch each bucket once, then look up matches in returned events
fetched_events = {}  # (sport_key, ym) -> [events]
for (sk, yr, mo), bucket_keys in buckets.items():
    # query the whole month, more efficient than per-day
    from calendar import monthrange
    last_day = monthrange(yr, mo)[1]
    frm = f"{yr:04d}-{mo:02d}-01T00:00:00Z"
    to  = f"{yr:04d}-{mo:02d}-{last_day:02d}T23:59:59Z"
    try:
        r = requests.get("https://api2.odds-api.io/v3/historical/events", params={
            "apiKey": API_KEY, "sport": "football", "league": sk,
            "from": frm, "to": to,
        }, timeout=20)
        if r.status_code != 200:
            print(f"  WARN: {sk} {yr}-{mo:02d} HTTP {r.status_code}")
            fetched_events[(sk, yr, mo)] = []
            continue
        data = r.json()
        events = data.get("data", data) if isinstance(data, dict) else data
        fetched_events[(sk, yr, mo)] = events if isinstance(events, list) else []
    except Exception as e:
        print(f"  WARN: {sk} {yr}-{mo:02d} {e}")
        fetched_events[(sk, yr, mo)] = []
    time.sleep(0.15)

# 4. Compare each tracker match to API result
mismatches = []
no_api_match = []
for key in keys_sorted:
    d, h, a, comp = key
    sk = ODDS_API_LEAGUES.get(comp)
    if not sk: continue
    events = fetched_events.get((sk, d.year, d.month), [])
    # Find matching event by date + teams
    found = None
    for ev in events:
        ed = ev.get("date", "")
        try:
            edate = datetime.fromisoformat(ed.replace("Z", "+00:00")).date()
        except Exception:
            continue
        # Tolerate ±1 day for TZ
        if abs((edate - d).days) > 1: continue
        if str(ev.get("home", "")).lower() == h.lower() and str(ev.get("away", "")).lower() == a.lower():
            found = ev
            break
    if not found:
        no_api_match.append(key)
        continue
    if found.get("status") != "settled":
        continue
    sc = found.get("scores", {})
    new_h, new_a, basis = derive(sc)
    stored = (matches[key]["hg"], matches[key]["ag"])
    if new_h is None or new_a is None:
        continue
    if (new_h, new_a) != stored:
        mismatches.append({
            "date": d, "home": h, "away": a, "comp": comp,
            "stored": stored, "api": (new_h, new_a),
            "basis": basis, "rows": matches[key]["rows"],
            "raw_scores": sc,
        })

# 5. Report
print(f"\n=== AUDIT RESULT ===")
print(f"Matches audited: {len(keys_sorted) - len(skipped_unmapped)}")
print(f"Matches not found in API response: {len(no_api_match)}")
print(f"Mismatches (stored != re-derived): {len(mismatches)}")
print()
for m in mismatches:
    print(f"  {m['date']} {m['home']} vs {m['away']} ({m['comp']})")
    print(f"     stored: {m['stored']}    re-derived: {m['api']}   basis: {m['basis']}")
    print(f"     raw scores: {m['raw_scores']}")
    print(f"     rows affected: {m['rows']}")
    print()

# Persist for follow-up fix
import json
out = Path("/tmp/audit_results_mismatches.json")
out.write_text(json.dumps(
    [{**m, "date": str(m["date"]), "raw_scores": m["raw_scores"]} for m in mismatches],
    indent=2, default=str))
print(f"Mismatch details saved to {out}")
