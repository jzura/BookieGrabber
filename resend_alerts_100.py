"""One-off: resend today's Telegram alerts (from 03:00 AWST) at €100/unit.

Reads the Master Bet Tracker xlsx, reapplies qualification logic per batch
(grouped by odds_time), and resends via telegram_alerts.send_bet_alerts.

Requires STAKE_PER_UNIT_WEEKEND to be temporarily set to 100 in
strategy_config.py and the telegrambot launchd process restarted.
"""

import sys
import yaml
import logging
from collections import defaultdict
from datetime import date, datetime, time as dt_time
from openpyxl import load_workbook

sys.path.insert(0, "/Users/Joel/REPOS/BookieGrabber")

from strategy_config import (
    compute_rpd, is_core_qualifying, is_btts_fade, is_15g_fade,
    DOUBLE_STAKE_RPD, DOUBLE_STAKE_MIN_COUNT, G15_FADE_RPD,
    STAKE_PER_UNIT_WEEKEND,
)
from telegram_alerts import send_bet_alerts
import json
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("resend")

MASTER_PATH = "/Users/Joel/Desktop/EFB_Master_Bet_Tracker_VS Code.xlsx"
SM_EVENT_MAP = Path("/Users/Joel/REPOS/BookieGrabber/data/state/sm_event_map.json")
CONFIG_YAML = "/Users/Joel/REPOS/BookieGrabber/config.yaml"
TODAY = date(2026, 5, 17)
FIRST_ROW_TODAY = 20830  # confirmed: rows 20820-29 are yesterday's late recordings

# Build competition→slug map (same as bet_tracker_updater)
with open(CONFIG_YAML) as f:
    cfg = yaml.safe_load(f)
comp_to_slug = {}
for lg in cfg.get("leagues", []):
    comp_to_slug[lg["name"]] = lg["slug"]
    comp_to_slug[lg["slug"]] = lg["slug"]
comp_to_slug.update({
    "Danish Superligaen": "danish_superliga",
    "Belgian First Division A": "belgian_pro_league",
    "Czech First League": "czech_1_liga",
    "English Championship League": "english_sky_bet_championship",
    "German Bundesliga I": "german_bundesliga",
    "Serbian SuperLiga": "serbian_super_league",
    "Sweden - Allsvenskan": "swedish_allsvenskan",
    "Turkish Super Lig": "turkish_super_league",
    "Netherlands Eredivisie": "dutch_eredivisie",
})

# Load preflight SM event map
sm_map = json.loads(SM_EVENT_MAP.read_text())


def describe_bet(bt, pred, rpd):
    is_btts_fade_b = (bt == "BTTS" and pred == 0 and rpd is not None and rpd >= 5)
    is_15g_fade_b = (bt == "1.5G" and pred == 1 and rpd is not None and rpd >= G15_FADE_RPD)
    if bt == "BTTS":
        if is_btts_fade_b:
            return "Bet BTTS Yes (No fade)"
        return "Bet BTTS No"
    if is_15g_fade_b:
        return "Bet 1.5G Under (fade)"
    side = "Over" if pred == 1 else "Under"
    return f"Bet {bt} {side}"


def actual_prediction(bt, pred, rpd):
    if bt == "BTTS" and pred == 0 and rpd is not None and rpd >= 5:
        return 1
    if bt == "1.5G" and pred == 1 and rpd is not None and rpd >= G15_FADE_RPD:
        return 0
    return pred


def main():
    assert STAKE_PER_UNIT_WEEKEND == 100.0, (
        f"Expected STAKE_PER_UNIT_WEEKEND=100, got {STAKE_PER_UNIT_WEEKEND}. "
        "Edit strategy_config.py first."
    )

    wb = load_workbook(MASTER_PATH, read_only=True)
    ws = wb.active

    # Collect today's rows from FIRST_ROW_TODAY onward
    rows = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=FIRST_ROW_TODAY, values_only=True), start=FIRST_ROW_TODAY):
        if not row or row[0] is None:
            continue
        bt, dt, home, away, comp, mt, ot, pred, o365, bf, vol = row[:11]
        if not isinstance(dt, datetime) or dt.date() < TODAY:
            continue
        rows.append({
            "row": r_idx, "bet_type": bt, "date": dt.date(),
            "home_team": home, "away_team": away, "competition": comp,
            "match_time": mt.strftime("%H:%M") if isinstance(mt, dt_time) else str(mt) if mt else "",
            "odds_time": ot.strftime("%H:%M") if isinstance(ot, dt_time) else str(ot) if ot else "",
            "prediction": pred, "odds_365": o365, "bf": bf, "volume": vol,
        })
    wb.close()
    log.info(f"Loaded {len(rows)} candidate rows from master (rows {FIRST_ROW_TODAY}+)")

    # Group by batch = odds_time (each pipeline run shared one odds_time)
    batches = defaultdict(list)
    for r in rows:
        batches[r["odds_time"]].append(r)
    log.info(f"Identified {len(batches)} batches by odds_time: "
             f"{sorted(batches.keys())}")

    alert_bets = []
    for ot, batch_rows in sorted(batches.items()):
        # Within batch: qualify, double-stake logic
        core_idx = set()
        match_core = defaultdict(int)
        for i, rd in enumerate(batch_rows):
            rpd = compute_rpd(rd["odds_365"], rd["bf"])
            rd["rpd"] = rpd
            if is_core_qualifying(rd["bet_type"], rd["prediction"], rd["bf"], rd["volume"], rpd):
                core_idx.add(i)
                mk = (str(rd["date"]), rd["home_team"], rd["away_team"])
                match_core[mk] += 1

        # Double stake: per match, only the highest-BF core row gets 2x
        match_dbl = defaultdict(list)
        for i in core_idx:
            rd = batch_rows[i]
            if rd["rpd"] is not None and rd["rpd"] <= DOUBLE_STAKE_RPD:
                mk = (str(rd["date"]), rd["home_team"], rd["away_team"])
                if match_core[mk] >= DOUBLE_STAKE_MIN_COUNT:
                    match_dbl[mk].append((i, float(rd["bf"] or 0)))
        double_idx = set()
        for mk, cands in match_dbl.items():
            best = max(cands, key=lambda x: x[1])[0]
            double_idx.add(best)

        for i, rd in enumerate(batch_rows):
            stake = None
            if i in core_idx:
                stake = 2 if i in double_idx else 1
            elif is_btts_fade(rd["bet_type"], rd["prediction"], rd["rpd"], rd["volume"]):
                stake = 1
            elif is_15g_fade(rd["bet_type"], rd["prediction"], rd["rpd"], rd["volume"]):
                stake = 1
            if not stake:
                continue

            pred = rd["prediction"]
            rpd = rd["rpd"]
            bt = rd["bet_type"]
            slug = comp_to_slug.get(rd["competition"], "")

            key = f"{rd['home_team']}|{rd['away_team']}"
            sm_entry = sm_map.get(key)
            if sm_entry:
                sm_event_id = sm_entry["sm_event_id"]
            else:
                # Synthetic ID — non-comma so bot falls back to live SM lookup
                # via _find_sm_event_id at button-press time.
                import hashlib
                h = hashlib.md5(f"{rd['date']}|{key}".encode()).hexdigest()[:10]
                sm_event_id = f"resend-{h}"

            alert_bets.append({
                "bet_type": bt,
                "home_team": rd["home_team"],
                "away_team": rd["away_team"],
                "competition": rd["competition"],
                "bf": rd["bf"] or 0,
                "rpd": rpd or 0,
                "volume": rd["volume"] or 0,
                "stake": stake,
                "match_time": rd["match_time"],
                "odds_time": rd["odds_time"],
                "date": str(rd["date"]),
                "prediction": pred,
                "actual_prediction": actual_prediction(bt, pred, rpd),
                "description": describe_bet(bt, pred, rpd),
                "event_id": sm_event_id,
                "league_slug": slug,
                "_batch": ot,
                "_sm_resolved": bool(sm_event_id),
            })

    log.info(f"Found {len(alert_bets)} qualifying bets to resend:")
    for b in alert_bets:
        marker = "OK " if b["_sm_resolved"] else "NO-SM"
        slug = b["league_slug"] or "NO-SLUG"
        log.info(
            f"  [{b['_batch']}] {marker} {slug:30} {b['description']:30} "
            f"{b['home_team']} vs {b['away_team']} | "
            f"stake={b['stake']} BF={b['bf']:.2f} RPD={b['rpd']:.2f} KO={b['match_time']}"
        )

    # Summarise
    by_batch = defaultdict(int)
    unresolved = 0
    for b in alert_bets:
        by_batch[b["_batch"]] += 1
        if not b["_sm_resolved"]:
            unresolved += 1
    log.info(f"\nBy batch: {dict(by_batch)}")
    log.info(f"Unresolved SM event IDs: {unresolved} (these will fail at button press)")

    if "--send" not in sys.argv:
        log.info("\nDRY RUN — pass --send to actually fire Telegram messages.")
        return

    # Clean up internal fields before sending
    for b in alert_bets:
        b.pop("_batch", None)
        b.pop("_sm_resolved", None)

    log.info(f"\nSENDING {len(alert_bets)} alerts via Telegram...")
    send_bet_alerts(alert_bets)
    log.info("Done.")


if __name__ == "__main__":
    main()
