"""
Automated Bet Tracker Update System

Processes ready_games DataFrames (totals + BTTS) and appends transformed data
to the master bet tracker spreadsheet on the desktop.

Can be called directly from the bookie_grabber pipeline, or used standalone
to process a local .xlsx file.

Usage:
    # Standalone: process a local file
    python bet_tracker_updater.py --local data/ready/ready_games_xxx.xlsx

    # From pipeline: called automatically by bookie_grabber.py
"""

import os
import sys
import json
import logging
import smtplib
import argparse
from email.message import EmailMessage
from pathlib import Path
from datetime import datetime, time as dt_time

import pandas as pd
from copy import copy
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

# -------------------------------------------------------------
# Configuration
# -------------------------------------------------------------

MASTER_PATH = Path.home() / "Desktop" / "EFB_Master_Bet_Tracker_VS Code.xlsx"
PENDING_RETRIES_PATH = Path(__file__).resolve().parent / "pending_retries.json"
FAILED_WRITES_PATH = Path(__file__).resolve().parent / "failed_master_writes.json"
MASTER_SHEET = "Master Bet Tracker"

logger = logging.getLogger("bookie_grabber.tracker_updater")

LINE_TO_BET_TYPE = {
    "Over/Under 1.5 Goals": "1.5G",
    "Over/Under 2.5 Goals": "2.5G",
    "Over/Under 3.5 Goals": "3.5G",
}

# Canonical league names for the master spreadsheet
LEAGUE_NAME_MAP = {
    "England - Premier League": "English Premier League",
    "England - Championship": "English Championship League",
    "International Clubs - UEFA Champions League": "UEFA Champions League",
    "Serbia - Superliga": "Serbian SuperLiga",
    "Portugal - Liga Portugal": "Portuguese Primeira Liga",
    "Germany - Bundesliga": "German Bundesliga I",
    "Czechia - 1. Liga": "Czech First League",
    "Romania - Superliga": "Romanian Liga I",
    "Turkiye - Super Lig": "Turkish Super Lig",
    "Scotland - Premiership": "Scottish Premiership",
    "Spain - LaLiga": "Spanish La Liga",
    "Netherlands - Eredivisie": "Netherlands Eredivisie",
    "Belgium - Pro League": "Belgian First Division A",
    "Croatia - HNL": "Croatian HNL",
    "France - Ligue 1": "French Ligue 1",
    "Poland - Ekstraklasa": "Polish Ekstraklasa",
    "Italy - Serie A": "Italian Serie A",
    "Switzerland - Super League": "Swiss Super League",
    "Serbian Super Liga": "Serbian SuperLiga",
    "Austria - Bundesliga": "Austrian Bundesliga",
    "Denmark - Superliga": "Danish Superligaen",
    "Greece - Super League": "Greek Super League",
    "Norway - Eliteserien": "Norwegian Eliteserien",
}


def normalize_league_name(name):
    """Map Odds API league names to canonical master spreadsheet names."""
    return LEAGUE_NAME_MAP.get(name, name)


# -------------------------------------------------------------
# Pending retries (for rows with missing Bet365 odds)
# -------------------------------------------------------------

def _track_missing_odds_row(row, bet_type, market):
    """Called from transform_totals/transform_btts when a row has NaN RPD
    (meaning Bet365 odds are missing). Record it so retry_missing_odds.py
    can re-fetch the odds and insert the row later."""
    import json as _json
    event_id = row.get("event_id")
    if not event_id:
        return

    # Parse match date/time
    mt = row.get("match_time")
    match_date_iso = None
    match_time_hm = None
    try:
        if isinstance(mt, str):
            mt = pd.to_datetime(mt)
        if hasattr(mt, "date"):
            match_date_iso = mt.date().isoformat()
            match_time_hm = mt.strftime("%H:%M")
    except Exception:
        pass

    pending = {}
    if PENDING_RETRIES_PATH.exists():
        try:
            with open(PENDING_RETRIES_PATH, "r") as f:
                pending = _json.load(f)
        except Exception:
            pending = {}

    key = str(event_id)
    entry = pending.get(key, {
        "event_id": int(event_id),
        "home_team": row.get("home_team"),
        "away_team": row.get("away_team"),
        "competition": normalize_league_name(row.get("competition", "")),
        "date": match_date_iso,
        "match_time": match_time_hm,
        "missing_bet_types": [],
        "added_at": datetime.now().isoformat(),
        "attempts": 0,
    })
    if bet_type not in entry["missing_bet_types"]:
        entry["missing_bet_types"].append(bet_type)
    pending[key] = entry

    try:
        with open(PENDING_RETRIES_PATH, "w") as f:
            _json.dump(pending, f, indent=2, default=str)
    except Exception:
        logger.exception("Failed to write pending_retries.json")


def _track_pending_retries(new_rows, start_row):
    """Record any rows with missing Bet365 odds so retry_missing_odds.py can
    re-fetch them later. Keyed by event_id so duplicate rows for the same
    match get consolidated."""
    import json as _json
    pending = {}
    if PENDING_RETRIES_PATH.exists():
        try:
            with open(PENDING_RETRIES_PATH, "r") as f:
                pending = _json.load(f)
        except Exception:
            pending = {}

    added = 0
    for i, rd in enumerate(new_rows):
        odds_365 = rd.get("odds_365")
        if odds_365 is not None and not pd.isna(odds_365):
            continue  # odds are present, nothing to retry

        event_id = rd.get("event_id")
        if not event_id:
            continue

        key = str(event_id)
        entry = pending.get(key, {
            "event_id": event_id,
            "home_team": rd.get("home_team"),
            "away_team": rd.get("away_team"),
            "competition": rd.get("competition"),
            "date": str(rd.get("date")) if rd.get("date") else None,
            "match_time": rd.get("match_time"),
            "missing_bet_types": [],
            "added_at": datetime.now().isoformat(),
            "attempts": 0,
        })
        if rd["bet_type"] not in entry["missing_bet_types"]:
            entry["missing_bet_types"].append(rd["bet_type"])
        pending[key] = entry
        added += 1

    if added:
        try:
            with open(PENDING_RETRIES_PATH, "w") as f:
                _json.dump(pending, f, indent=2, default=str)
            logger.info(f"Tracked {added} row(s) with missing Bet365 odds for retry")
        except Exception:
            logger.exception("Failed to write pending_retries.json")


# -------------------------------------------------------------
# Stake alert email
# -------------------------------------------------------------

def _describe_bet(row_data):
    """Return a human-readable bet description showing the ACTUAL bet to place.
    For fade bets, the bet is the OPPOSITE of the prediction."""
    bt = row_data["bet_type"]
    pred = row_data["prediction"]
    rpd = row_data.get("rpd")

    is_btts_fade = (bt == "BTTS" and pred == 0 and rpd is not None and rpd >= 5)
    is_1_5g_fade = (bt == "1.5G" and pred == 1 and rpd is not None and rpd >= 4.6)

    if bt == "BTTS":
        if is_btts_fade:
            return "Bet BTTS Yes (No fade)"
        return "Bet BTTS No"

    if is_1_5g_fade:
        return "Bet 1.5G Under (fade)"

    side = "Over" if pred == 1 else "Under"
    return f"Bet {bt} {side}"


def send_stake_alerts(bets):
    """
    Send an email listing bets that qualify for a stake.
    bets: list of dicts with keys: competition, bet_type, prediction, stake,
          home_team, away_team, match_time, bf, rpd
    """
    if not bets:
        return

    email_user = os.environ.get("EMAIL_USER")
    email_pass = os.environ.get("EMAIL_PASS")
    email_to = os.environ.get("EMAIL_TO", "")
    recipients = [e.strip() for e in email_to.split(",") if e.strip()]

    if not email_user or not email_pass or not recipients:
        logger.warning("Email not configured — skipping stake alerts")
        return

    lines = []
    for b in bets:
        desc = _describe_bet(b)
        stake_str = f"{b['stake']}x" if b["stake"] == 2 else "1x"
        lines.append(
            f"  {stake_str}  {b['competition']}  |  "
            f"{b['home_team']} vs {b['away_team']}  |  "
            f"{desc}  |  BF {b['bf']}  |  KO {b['match_time']}"
        )

    body = f"{len(bets)} bet(s) to place:\n\n" + "\n".join(lines)

    msg = EmailMessage()
    msg["From"] = email_user
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = f"Stake Alert — {len(bets)} bet(s)"
    msg.set_content(body)

    try:
        with smtplib.SMTP("smtp.mail.me.com", 587, timeout=15) as server:
            server.starttls()
            server.login(email_user, email_pass)
            server.send_message(msg)
        logger.info(f"Stake alert email sent: {len(bets)} bet(s)")
    except Exception as e:
        logger.exception(f"Failed to send stake alert: {e}")

    # Also send via Telegram
    try:
        from telegram_alerts import send_bet_alerts as tg_send
        # Build competition name → config slug lookup for SM URL routing
        # Master sheet comp names often differ from config names, so build
        # a mapping that handles common variations
        import yaml
        with open("config.yaml") as _f:
            _cfg = yaml.safe_load(_f)
        _comp_to_slug = {}
        for lg in _cfg.get("leagues", []):
            _comp_to_slug[lg["name"]] = lg["slug"]
            # Also add slug itself and common variations
            _comp_to_slug[lg["slug"]] = lg["slug"]
        # Add known master-sheet-to-config name mappings
        _comp_aliases = {
            "Danish Superligaen": "danish_superliga",
            "Belgian First Division A": "belgian_pro_league",
            "Czech First League": "czech_1_liga",
            "English Championship League": "english_sky_bet_championship",
            "German Bundesliga I": "german_bundesliga",
            "Serbian SuperLiga": "serbian_super_league",
            "Sweden - Allsvenskan": "swedish_allsvenskan",
            "Turkish Super Lig": "turkish_super_league",
        }
        _comp_to_slug.update(_comp_aliases)

        tg_bets = []
        for b in bets:
            desc = _describe_bet(b)
            rpd_val = b.get("rpd")
            if rpd_val is None:
                rpd_val = _compute_rpd(b.get("odds_365"), b.get("bf"))
            # Resolve the actual prediction to place (fades flip the side)
            bt = b["bet_type"]
            pred = b["prediction"]
            rpd = b.get("rpd")
            is_btts_fade = (bt == "BTTS" and pred == 0 and rpd is not None and rpd >= 5)
            is_1_5g_fade = (bt == "1.5G" and pred == 1 and rpd is not None and rpd >= 4.6)
            if is_btts_fade:
                actual_pred = 1  # Bet Yes instead of No
            elif is_1_5g_fade:
                actual_pred = 0  # Bet Under instead of Over
            else:
                actual_pred = pred
            league_slug = _comp_to_slug.get(b.get("competition"), "")
            if not league_slug:
                logger.warning(f"No league_slug for competition '{b.get('competition')}' — SM placement will fail")
            tg_bets.append({
                **b,
                "description": desc,
                "rpd": rpd_val or 0,
                "actual_prediction": actual_pred,
                "league_slug": league_slug,
            })

        # Resolve SM event IDs for all bets in one Playwright session
        _resolve_sm_event_ids(tg_bets)

        tg_send(tg_bets)
        logger.info(f"Telegram alert sent: {len(bets)} bet(s)")
    except Exception as e:
        logger.warning(f"Telegram alert failed: {e}")


def _resolve_sm_event_ids(tg_bets):
    """Resolve SportsMarket event IDs for a list of Telegram bets.

    Opens one Playwright session, visits each league's SM sportsbook page,
    extracts all event links, and matches them to our bets by team name.
    Modifies each bet dict in-place to set 'event_id' to the SM format
    (e.g. '2026-04-18,994,996').
    """
    from sportsmarket_api import SM_LEAGUE_PREFIXES, SM_SPORTSBOOK_BASE
    from difflib import SequenceMatcher
    import re

    # Group bets by league_slug to minimize page loads
    from collections import defaultdict
    bets_by_league = defaultdict(list)
    for b in tg_bets:
        slug = b.get("league_slug", "")
        if slug:
            bets_by_league[slug].append(b)

    if not bets_by_league:
        return

    try:
        from playwright.sync_api import sync_playwright
        password = os.environ.get("SM_PASSWORD", "")
        username = os.environ.get("SM_USERNAME", "joelbrown95")
        if not password:
            logger.warning("SM_PASSWORD not set — cannot resolve SM event IDs")
            return

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            ctx = browser.new_context(viewport={"width": 1400, "height": 900})
            page = ctx.new_page()

            # Login
            page.goto("https://pro.sportmarket.com/login", timeout=30000)
            page.wait_for_selector('input[type="text"]', timeout=10000)
            page.fill('input[type="text"]', username)
            page.fill('input[type="password"]', password)
            page.click('button[data-testid="35eb9af8"]')
            page.wait_for_url(
                lambda url: "/sportsbook" in url or "/trade" in url,
                timeout=30000)

            for slug, bets in bets_by_league.items():
                prefix = SM_LEAGUE_PREFIXES.get(slug)
                if not prefix:
                    logger.warning(f"No SM prefix for league {slug}")
                    continue

                page.goto(f"{SM_SPORTSBOOK_BASE}/{prefix}", timeout=15000)
                page.wait_for_load_state("networkidle", timeout=10000)
                page.wait_for_timeout(3000)

                # Extract all event links on the page (any link with date,id,id pattern)
                sm_events = page.evaluate('''() => {
                    const links = document.querySelectorAll('a[href*=","]');
                    let r = [];
                    links.forEach(l => {
                        const href = l.getAttribute('href') || '';
                        const match = href.match(/(\\d{4}-\\d{2}-\\d{2},\\d+,\\d+)/);
                        if (match && href.includes('/sportsbook/')) {
                            r.push({eid: match[1], text: l.textContent.trim()});
                        }
                    });
                    return r;
                }''')

                # Match each bet to an SM event by team name similarity
                for b in bets:
                    home = b.get("home_team", "")
                    away = b.get("away_team", "")
                    best_score = 0
                    best_eid = None
                    for ev in sm_events:
                        txt = ev["text"].lower()
                        h_score = SequenceMatcher(None, home.lower(), txt).ratio()
                        a_score = SequenceMatcher(None, away.lower(), txt).ratio()
                        # Both teams should appear in the event text
                        score = h_score + a_score
                        if score > best_score:
                            best_score = score
                            best_eid = ev["eid"]
                    if best_eid and best_score > 0.6:
                        b["event_id"] = best_eid
                        logger.info(f"Resolved SM event: {home} vs {away} -> {best_eid}")
                    else:
                        logger.warning(f"Could not resolve SM event for {home} vs {away} "
                                      f"(best_score={best_score:.2f})")

            browser.close()
    except Exception as e:
        logger.warning(f"SM event ID resolution failed: {e}")


# -------------------------------------------------------------
# Data transformation
# -------------------------------------------------------------

def parse_datetime_parts(dt_value):
    """Parse a datetime value into (date, HH:MM string)."""
    if pd.isna(dt_value):
        return None, None
    if isinstance(dt_value, str):
        dt_value = pd.to_datetime(dt_value)
    if isinstance(dt_value, (datetime, pd.Timestamp)):
        return dt_value.date(), dt_value.strftime("%H:%M")
    return None, None


def transform_totals(df):
    """Transform totals_ready DataFrame into master sheet rows."""
    rows = []
    for _, row in df.iterrows():
        line = row.get("line", "")
        bet_type = LINE_TO_BET_TYPE.get(line)
        if not bet_type:
            logger.warning(f"Unknown line value: {line}, skipping row")
            continue

        over_rpd = pd.to_numeric(row.get("Over RPD"), errors="coerce")
        under_rpd = pd.to_numeric(row.get("Under RPD"), errors="coerce")
        if pd.isna(over_rpd) or pd.isna(under_rpd):
            # Track for later retry instead of silently dropping
            _track_missing_odds_row(row, bet_type, market="totals")
            logger.warning(f"Missing RPD for {row.get('home_team')} vs {row.get('away_team')} ({bet_type}), tracked for retry")
            continue

        # Determine prediction: lower RPD wins
        if under_rpd < over_rpd:
            prediction = 0  # Under
        else:
            prediction = 1  # Over

        # 1.5G exception: Over RPD >= 4.6 forces Under
        if bet_type == "1.5G" and over_rpd >= 4.6:
            prediction = 0

        # Select odds for predicted side
        if prediction == 1:
            odds_365 = pd.to_numeric(row.get("Bet365_over_odds"), errors="coerce")
            bf_odds = pd.to_numeric(row.get("Betfair_Exchange_over_odds"), errors="coerce")
            rpd = over_rpd
        else:
            odds_365 = pd.to_numeric(row.get("Bet365_under_odds"), errors="coerce")
            bf_odds = pd.to_numeric(row.get("Betfair_Exchange_under_odds"), errors="coerce")
            rpd = under_rpd

        match_date, match_time = parse_datetime_parts(row.get("match_time"))
        _, odds_time = parse_datetime_parts(row.get("odds_time"))
        volume = pd.to_numeric(row.get("total_volume"), errors="coerce")

        rows.append({
            "bet_type": bet_type,
            "date": match_date,
            "home_team": row.get("home_team"),
            "away_team": row.get("away_team"),
            "competition": normalize_league_name(row.get("competition", "")),
            "match_time": match_time,
            "odds_time": odds_time,
            "prediction": prediction,
            "odds_365": odds_365,
            "bf": bf_odds,
            "volume": round(volume, 2) if not pd.isna(volume) else None,
            "rpd": round(rpd, 3),
            "event_id": row.get("event_id"),
        })
    return rows


def transform_btts(df):
    """Transform btts_ready DataFrame into master sheet rows."""
    rows = []
    for _, row in df.iterrows():
        yes_rpd = pd.to_numeric(row.get("Yes RPD"), errors="coerce")
        no_rpd = pd.to_numeric(row.get("No RPD"), errors="coerce")
        if pd.isna(yes_rpd) or pd.isna(no_rpd):
            _track_missing_odds_row(row, "BTTS", market="btts")
            logger.warning(f"Missing RPD for BTTS {row.get('home_team')} vs {row.get('away_team')}, tracked for retry")
            continue

        if no_rpd < yes_rpd:
            prediction = 0  # No
        else:
            prediction = 1  # Yes

        if prediction == 1:
            odds_365 = pd.to_numeric(row.get("Bet365_yes_odds"), errors="coerce")
            bf_odds = pd.to_numeric(row.get("Betfair_Exchange_yes_odds"), errors="coerce")
            rpd = yes_rpd
        else:
            odds_365 = pd.to_numeric(row.get("Bet365_no_odds"), errors="coerce")
            bf_odds = pd.to_numeric(row.get("Betfair_Exchange_no_odds"), errors="coerce")
            rpd = no_rpd

        match_date, match_time = parse_datetime_parts(row.get("match_time"))
        _, odds_time = parse_datetime_parts(row.get("odds_time"))
        volume = pd.to_numeric(row.get("total_volume"), errors="coerce")

        rows.append({
            "bet_type": "BTTS",
            "date": match_date,
            "home_team": row.get("home_team"),
            "away_team": row.get("away_team"),
            "competition": normalize_league_name(row.get("competition", "")),
            "match_time": match_time,
            "odds_time": odds_time,
            "prediction": prediction,
            "odds_365": odds_365,
            "bf": bf_odds,
            "volume": round(volume, 2) if not pd.isna(volume) else None,
            "rpd": round(rpd, 3),
            "rpd_yes": round(float(yes_rpd), 3) if not pd.isna(yes_rpd) else None,
            "rpd_no": round(float(no_rpd), 3) if not pd.isna(no_rpd) else None,
            "event_id": row.get("event_id"),
        })
    return rows


# -------------------------------------------------------------
# Duplicate detection
# -------------------------------------------------------------

def load_existing_keys(ws):
    """Load (bet_type, date, home_team, away_team) from master sheet for dedup."""
    keys = set()
    for row_num in range(2, ws.max_row + 1):
        bet_type = ws.cell(row=row_num, column=1).value
        date_val = ws.cell(row=row_num, column=2).value
        home = ws.cell(row=row_num, column=3).value
        away = ws.cell(row=row_num, column=4).value
        if bet_type is None:
            continue
        if isinstance(date_val, datetime):
            date_val = date_val.date()
        keys.add((str(bet_type), str(date_val), str(home), str(away)))
    return keys


# -------------------------------------------------------------
# Cross-market conflict detection
# -------------------------------------------------------------

def load_existing_predictions(ws):
    """Load (bet_type, date, home, away) -> prediction from master sheet."""
    predictions = {}
    for row_num in range(2, ws.max_row + 1):
        bet_type = ws.cell(row=row_num, column=1).value
        date_val = ws.cell(row=row_num, column=2).value
        home = ws.cell(row=row_num, column=3).value
        away = ws.cell(row=row_num, column=4).value
        pred = ws.cell(row=row_num, column=8).value
        if bet_type is None:
            continue
        if isinstance(date_val, datetime):
            date_val = date_val.date()
        predictions[(str(bet_type), str(date_val), str(home), str(away))] = pred
    return predictions


def detect_conflicts(all_rows, existing_predictions):
    """
    Detect cross-market conflicts: NONE currently active.
    Both BTTS and 3.5G conflict filters have been removed.
    Returns empty set.
    """
    return set()


# -------------------------------------------------------------
# Formula generation (matching existing master sheet patterns)
# -------------------------------------------------------------

def rpd_formula(r):
    return (
        f'=IF(OR(ISBLANK(I{r}),ISBLANK(J{r})),"",IF(I{r}>J{r},1,'
        f'IF(ABS(I{r}-J{r})/((I{r}+J{r})/2)*100<1,1,ABS(I{r}-J{r})/((I{r}+J{r})/2)*100)))'
    )

def result_formula(r):
    """Result formula. For BTTS, derived from N (home goals) and O (away goals).
    For goals markets, derived from M (total goals) and the line in column A."""
    return (
        f'=IF(A{r}="BTTS",'
        f'IF(OR(N{r}="",O{r}=""),"",IF(H{r}=1,IF(AND(N{r}>0,O{r}>0),1,0),IF(AND(N{r}>0,O{r}>0),0,1))),'
        f'IF(M{r}="","",IF(A{r}="1.5G",IF(OR(AND(M{r}>=2,H{r}=1),AND(M{r}<=1,H{r}=0)),1,0),'
        f'IF(A{r}="2.5G",IF(OR(AND(M{r}>=3,H{r}=1),AND(M{r}<=2,H{r}=0)),1,0),'
        f'IF(A{r}="3.5G",IF(OR(AND(M{r}>=4,H{r}=1),AND(M{r}<=3,H{r}=0)),1,0),"")))))'
    )

def goals_formula(r):
    """Total goals = home + away (auto-derived)."""
    return f'=IF(AND(N{r}="",O{r}=""),"",N{r}+O{r})'

def _core_conditions(r):
    """Core contrarian conditions: 1.5G/3.5G/BTTS, pred=0, volume+BF+RPD schedule."""
    from strategy_config import core_conditions_excel
    return core_conditions_excel(r)

def _fade_conditions(r):
    """Fade: BTTS pred=0 RPD>=5 (fade to Yes), 1.5G pred=1 RPD>=4.6 (fade to Under)."""
    from strategy_config import fade_conditions_excel
    return fade_conditions_excel(r)

def stake_formula(r, is_conflict=False, is_double_stake=False, is_25g_piggyback=False):
    if is_conflict:
        return '=""'
    if is_25g_piggyback:
        from strategy_config import piggyback_conditions_excel
        return f'=IF({piggyback_conditions_excel(r)},1,"")'
    if is_double_stake:
        return (
            f'=IF({_core_conditions(r)},2,IF(OR({_fade_conditions(r)}),1,""))'
        )
    return (
        f'=IF(OR({_core_conditions(r)},{_fade_conditions(r)}),1,"")'
    )

def return_formula(r):
    """SportsMarket tiered commission: 1% (<=1.5), 2% (<=2.8), 3% (<=3.5), 4% (>3.5).
    For fade bets, commission is based on the opposite odds 1/(1-1/J), not J.
    Q = Stake, P = Result, J = BF odds."""
    return (
        f'=IF(Q{r}="","",'
        f'IF(OR(AND(A{r}="BTTS",H{r}=0,L{r}>=5),AND(A{r}="1.5G",H{r}=1,L{r}>=4.6)),'
        f'IF(P{r}=0,'
        f'Q{r}*(1+(1/(1-1/J{r})-1)*(1-IF(1/(1-1/J{r})<=1.5,0.01,IF(1/(1-1/J{r})<=2.8,0.02,IF(1/(1-1/J{r})<=3.5,0.03,0.04))))),'
        f'0),'
        f'IF(P{r}=1,'
        f'Q{r}*(1+(J{r}-1)*(1-IF(J{r}<=1.5,0.01,IF(J{r}<=2.8,0.02,IF(J{r}<=3.5,0.03,0.04))))),'
        f'0)))'
    )

def profit_formula(r):
    return f'=IF(Q{r}="","",R{r}-Q{r})'

def cumulative_profit_formula(r):
    return f'=IF(S{r}="","",SUM(S$2:S{r}))'

def bet_number_formula(r):
    return f'=IF(Q{r}="","",COUNTIF(Q$2:Q{r},">0"))'


# -------------------------------------------------------------
# Master sheet updating
# -------------------------------------------------------------

def append_to_master(transformed_rows, master_path=MASTER_PATH):
    """
    Append transformed rows to the master bet tracker.
    Handles dedup, conflict detection, and formula insertion.
    Returns count of rows appended.
    """
    if not transformed_rows:
        logger.info("No rows to append")
        return 0

    if not master_path.exists():
        logger.error(f"Master spreadsheet not found: {master_path}")
        return 0

    wb = load_workbook(master_path)
    ws = wb[MASTER_SHEET]

    logger.info("Loading existing keys and predictions...")
    existing_keys = load_existing_keys(ws)
    existing_predictions = load_existing_predictions(ws)

    # Filter duplicates
    new_rows = []
    for r in transformed_rows:
        key = (str(r["bet_type"]), str(r["date"]), str(r["home_team"]), str(r["away_team"]))
        if key in existing_keys:
            logger.info(f"Skipping duplicate: {r['bet_type']} {r['date']} {r['home_team']} vs {r['away_team']}")
        else:
            new_rows.append(r)

    if not new_rows:
        logger.info("All rows are duplicates, nothing to append")
        wb.close()
        return 0

    # Detect conflicts
    conflict_indices = detect_conflicts(new_rows, existing_predictions)

    # Determine double-stake eligibility for new rows
    # A row gets 2x stake when: core qualifying + RPD=1.0 + >=2 core bets on same match
    def _compute_rpd(odds_365, bf):
        try:
            i_val, j_val = float(odds_365), float(bf)
            if i_val > j_val:
                return 1.0
            pct = abs(i_val - j_val) / ((i_val + j_val) / 2) * 100
            return 1.0 if pct < 1 else round(pct, 3)
        except (ValueError, TypeError, ZeroDivisionError):
            return None

    def _is_core_qualifying(row_data, idx):
        from strategy_config import is_core_qualifying as _is_core
        if idx in conflict_indices:
            return False
        rpd = _compute_rpd(row_data.get("odds_365"), row_data.get("bf"))
        return _is_core(row_data["bet_type"], row_data["prediction"],
                        row_data.get("bf"), row_data.get("volume"), rpd)

    from collections import defaultdict
    core_indices = set()
    match_core_count = defaultdict(int)
    for i, rd in enumerate(new_rows):
        if _is_core_qualifying(rd, i):
            core_indices.add(i)
            mk = (str(rd["date"]), str(rd["home_team"]), str(rd["away_team"]))
            match_core_count[mk] += 1

    # Double-stake: RPD=1.0 + ≥2 core bets on same match, but only highest BF per match
    match_dbl_candidates = defaultdict(list)
    for i in core_indices:
        rd = new_rows[i]
        rpd = _compute_rpd(rd.get("odds_365"), rd.get("bf"))
        mk = (str(rd["date"]), str(rd["home_team"]), str(rd["away_team"]))
        from strategy_config import DOUBLE_STAKE_RPD, DOUBLE_STAKE_MIN_COUNT
        if rpd is not None and rpd <= DOUBLE_STAKE_RPD and match_core_count[mk] >= DOUBLE_STAKE_MIN_COUNT:
            match_dbl_candidates[mk].append((i, float(rd.get("bf", 0))))

    double_stake_indices = set()
    for mk, candidates in match_dbl_candidates.items():
        best_i = max(candidates, key=lambda x: x[1])[0]
        double_stake_indices.add(best_i)

    # Under 2.5G piggyback: find matches where Under 1.5G qualifies as core,
    # then flag the corresponding Under 2.5G row on the same match
    matches_with_core_15g = set()
    for i in core_indices:
        rd = new_rows[i]
        if rd["bet_type"] == "1.5G" and rd["prediction"] == 0:
            mk = (str(rd["date"]), str(rd["home_team"]), str(rd["away_team"]))
            matches_with_core_15g.add(mk)

    piggyback_25g_indices = set()
    for i, rd in enumerate(new_rows):
        from strategy_config import is_25g_piggyback as _is_pb
        mk = (str(rd["date"]), str(rd["home_team"]), str(rd["away_team"]))
        if _is_pb(rd["bet_type"], rd["prediction"], rd.get("bf"), mk in matches_with_core_15g):
            piggyback_25g_indices.add(i)

    # Find actual last row with data in column A (ws.max_row can be inflated by empty formatted rows)
    last_data_row = 1  # header row
    for row_num in range(ws.max_row, 1, -1):
        if ws.cell(row=row_num, column=1).value is not None:
            last_data_row = row_num
            break

    # Capture formatting from the template row (last existing data row) for each column
    template_styles = {}
    for col in range(1, 22):  # A through U (Bet Number is now column U)
        src_cell = ws.cell(row=last_data_row, column=col)
        template_styles[col] = {
            "font": copy(src_cell.font),
            "fill": copy(src_cell.fill),
            "border": copy(src_cell.border),
            "alignment": copy(src_cell.alignment),
            "number_format": src_cell.number_format,
            "protection": copy(src_cell.protection),
        }

    def apply_template_style(cell, col):
        s = template_styles[col]
        cell.font = copy(s["font"])
        cell.fill = copy(s["fill"])
        cell.border = copy(s["border"])
        cell.alignment = copy(s["alignment"])
        cell.number_format = s["number_format"]
        cell.protection = copy(s["protection"])

    # Append rows
    start_row = last_data_row + 1
    appended = 0

    for i, row_data in enumerate(new_rows):
        r = start_row + i

        # Apply template formatting to all cells in this row first
        for col in range(1, 22):
            apply_template_style(ws.cell(row=r, column=col), col)

        ws.cell(row=r, column=1, value=row_data["bet_type"])

        if row_data["date"]:
            ws.cell(row=r, column=2, value=datetime.combine(row_data["date"], datetime.min.time()))

        ws.cell(row=r, column=3, value=row_data["home_team"])
        ws.cell(row=r, column=4, value=row_data["away_team"])
        ws.cell(row=r, column=5, value=row_data["competition"])

        if row_data["match_time"]:
            h, m = map(int, row_data["match_time"].split(":"))
            ws.cell(row=r, column=6, value=dt_time(h, m))

        if row_data["odds_time"]:
            h, m = map(int, row_data["odds_time"].split(":"))
            ws.cell(row=r, column=7, value=dt_time(h, m))

        ws.cell(row=r, column=8, value=row_data["prediction"])

        if not pd.isna(row_data["odds_365"]):
            ws.cell(row=r, column=9, value=row_data["odds_365"])

        if not pd.isna(row_data["bf"]):
            ws.cell(row=r, column=10, value=row_data["bf"])

        if row_data["volume"] is not None and not pd.isna(row_data["volume"]):
            ws.cell(row=r, column=11, value=row_data["volume"])

        # L: RPD formula
        ws.cell(row=r, column=12, value=rpd_formula(r))
        # M: Goals (formula =N+O, auto-derived from home/away)
        ws.cell(row=r, column=13, value=goals_formula(r))
        # N: Home Goals — blank (filled by results updater)
        # O: Away Goals — blank (filled by results updater)
        # P: Result formula
        ws.cell(row=r, column=16, value=result_formula(r))
        # Q: Stake formula
        ws.cell(row=r, column=17, value=stake_formula(
            r,
            is_conflict=(i in conflict_indices),
            is_double_stake=(i in double_stake_indices),
            is_25g_piggyback=(i in piggyback_25g_indices),
        ))
        # R: Return
        ws.cell(row=r, column=18, value=return_formula(r))
        # S: Profit
        ws.cell(row=r, column=19, value=profit_formula(r))
        # T: Cumulative Profit
        ws.cell(row=r, column=20, value=cumulative_profit_formula(r))
        # U: Bet Number
        ws.cell(row=r, column=21, value=bet_number_formula(r))

        appended += 1

    logger.info(f"Appending {appended} rows starting at row {start_row}")

    # Track rows with missing Bet365 odds for later retry
    _track_pending_retries(new_rows, start_row)

    # Load set of already-alerted bets to avoid duplicate alerts
    ALERTED_STATE_FILE = Path(__file__).resolve().parent / "data" / "state" / "alerted_bets.json"
    alerted_keys = set()
    if ALERTED_STATE_FILE.exists():
        try:
            all_keys = json.loads(ALERTED_STATE_FILE.read_text())
            # Prune keys older than 3 days (key format: "bt:date:home:away")
            today = date.today()
            for k in all_keys:
                try:
                    d = date.fromisoformat(k.split(":")[1])
                    if (today - d).days <= 3:
                        alerted_keys.add(k)
                except:
                    alerted_keys.add(k)  # keep unparseable keys
        except:
            pass

    # Collect qualifying bets for stake alert email
    alert_bets = []
    for i, row_data in enumerate(new_rows):
        if i in conflict_indices:
            continue

        bt = row_data["bet_type"]
        pred = row_data["prediction"]
        vol = row_data.get("volume")
        bf = row_data.get("bf")
        rpd = _compute_rpd(row_data.get("odds_365"), bf)

        stake = None

        from strategy_config import is_core_qualifying as _is_core, is_btts_fade as _is_bf, is_15g_fade as _is_1f

        # Core contrarian
        if _is_core(bt, pred, bf, vol, rpd):
            stake = 2 if i in double_stake_indices else 1

        # Fade: BTTS pred=0, RPD>=5 (fade to Yes)
        if _is_bf(bt, pred, rpd, vol):
            stake = stake or 1

        # Fade: 1.5G pred=1, RPD>=4.6 (fade to Under)
        if _is_1f(bt, pred, rpd, vol):
            stake = stake or 1

        # Under 2.5G piggyback (pre-qualified above)
        if i in piggyback_25g_indices:
            stake = stake or 1

        if stake:
            alert_key = f"{row_data['bet_type']}:{row_data['date']}:{row_data['home_team']}:{row_data['away_team']}"
            if alert_key in alerted_keys:
                continue  # already alerted — don't resend
            alert_bets.append({
                **row_data,
                "stake": stake,
            })
            alerted_keys.add(alert_key)

    if alert_bets:
        send_stake_alerts(alert_bets)
        # Persist alerted keys so we don't re-alert on next run
        ALERTED_STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
        ALERTED_STATE_FILE.write_text(json.dumps(list(alerted_keys), default=str))

    try:
        wb.save(master_path)
        wb.close()
        logger.info(f"Master spreadsheet saved: {master_path}")
    except PermissionError:
        wb.close()
        logger.warning(f"Master spreadsheet is locked (open in Excel?) — queuing {appended} rows for next run")
        _queue_failed_writes(new_rows)
        return 0
    except Exception as e:
        wb.close()
        logger.error(f"Failed to save master spreadsheet: {e} — queuing {appended} rows")
        _queue_failed_writes(new_rows)
        return 0

    return appended


# -------------------------------------------------------------
# Failed write queue
# -------------------------------------------------------------

def _queue_failed_writes(rows):
    """Save transformed rows to disk so they can be retried when the file is available."""
    import json as _json
    from datetime import datetime as _dt

    queued = []
    if FAILED_WRITES_PATH.exists():
        try:
            with open(FAILED_WRITES_PATH, "r") as f:
                queued = _json.load(f)
        except Exception:
            queued = []

    # Convert rows to serializable format
    for row in rows:
        entry = {}
        for k, v in row.items():
            if hasattr(v, "isoformat"):
                entry[k] = v.isoformat()
            elif isinstance(v, float) and pd.isna(v):
                entry[k] = None
            else:
                entry[k] = v
        entry["_queued_at"] = _dt.now().isoformat()
        queued.append(entry)

    try:
        with open(FAILED_WRITES_PATH, "w") as f:
            _json.dump(queued, f, indent=2, default=str)
        logger.info(f"Queued {len(rows)} rows to {FAILED_WRITES_PATH}")
    except Exception:
        logger.exception("Failed to write failed_master_writes.json")


def _retry_failed_writes(master_path=MASTER_PATH):
    """Attempt to append any queued rows from previous failed writes."""
    import json as _json

    if not FAILED_WRITES_PATH.exists():
        return 0

    try:
        with open(FAILED_WRITES_PATH, "r") as f:
            queued = _json.load(f)
    except Exception:
        return 0

    if not queued:
        return 0

    logger.info(f"Retrying {len(queued)} queued rows from failed writes")

    # Strip the _queued_at field and restore types
    rows = []
    for entry in queued:
        row = {k: v for k, v in entry.items() if k != "_queued_at"}
        # Restore match_time/odds_time to datetime if they're ISO strings
        for dt_field in ("match_time", "odds_time"):
            if dt_field in row and isinstance(row[dt_field], str):
                try:
                    row[dt_field] = pd.to_datetime(row[dt_field])
                except Exception:
                    pass
        rows.append(row)

    try:
        appended = append_to_master(rows, master_path)
        if appended > 0:
            # Clear the queue on success
            FAILED_WRITES_PATH.unlink(missing_ok=True)
            logger.info(f"Successfully retried {appended} queued rows")
        return appended
    except Exception:
        logger.exception("Retry of queued rows also failed — will try again next run")
        return 0


# -------------------------------------------------------------
# Public API for pipeline integration
# -------------------------------------------------------------

def update_master_from_dataframes(totals_df, btts_df, master_path=MASTER_PATH):
    """
    Called directly from bookie_grabber pipeline.
    Accepts the ready totals and BTTS DataFrames, transforms them,
    and appends to the master bet tracker.
    Returns number of rows appended.
    """
    # First, retry any previously failed writes
    retried = _retry_failed_writes(master_path)
    if retried > 0:
        logger.info(f"Retried {retried} previously queued rows")

    all_rows = []

    if totals_df is not None and not totals_df.empty:
        all_rows.extend(transform_totals(totals_df))
        logger.info(f"Transformed {len(all_rows)} totals rows")

    btts_start = len(all_rows)
    if btts_df is not None and not btts_df.empty:
        all_rows.extend(transform_btts(btts_df))
        logger.info(f"Transformed {len(all_rows) - btts_start} BTTS rows")

    if not all_rows:
        logger.info("No rows to update master with")
        return 0

    return append_to_master(all_rows, master_path)


def update_master_from_file(file_path):
    """Process a local ready_games .xlsx file and append to master sheet."""
    file_path = Path(file_path)
    logger.info(f"Processing file: {file_path}")

    totals_df = pd.DataFrame()
    btts_df = pd.DataFrame()

    try:
        df = pd.read_excel(file_path, sheet_name="totals_ready")
        if not df.empty and len(df.columns) > 1:
            totals_df = df
    except Exception as e:
        logger.warning(f"Could not read totals_ready: {e}")

    try:
        df = pd.read_excel(file_path, sheet_name="btts_ready")
        if not df.empty and len(df.columns) > 1:
            btts_df = df
    except Exception as e:
        logger.warning(f"Could not read btts_ready: {e}")

    return update_master_from_dataframes(totals_df, btts_df)


# -------------------------------------------------------------
# CLI for standalone use
# -------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Bet Tracker Updater — process ready_games into master sheet")
    parser.add_argument("--local", type=str, help="Process a local .xlsx file")
    args = parser.parse_args()

    # Set up logging for standalone use
    if not logger.handlers:
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s [%(levelname)s] %(message)s",
            handlers=[logging.StreamHandler()],
        )

    if args.local:
        local_path = Path(args.local)
        if not local_path.exists():
            logger.error(f"File not found: {local_path}")
            sys.exit(1)
        appended = update_master_from_file(local_path)
        logger.info(f"Done. {appended} rows appended to master sheet.")
    else:
        print("Usage: python bet_tracker_updater.py --local <path_to_ready_games.xlsx>")
        print("Or import and call update_master_from_dataframes() from the pipeline.")


if __name__ == "__main__":
    main()