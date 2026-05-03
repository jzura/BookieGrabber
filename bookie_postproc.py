"""
PRODUCTION-READY bookie_grabber
Generates:
    - master per-league CSVs (daily)
    - per-day Excel workbook with all "ready" matches (totals + btts + merged)
    - processed_cache.json to avoid duplicate exports across multiple runs

Excludes: email module and scheduler (you said you'll use launchd on macOS).

Requirements:
    pip install -r requirements.txt
    (requests, pyyaml, python-dotenv, pandas, openpyxl, pytz, python-dateutil)

Drop into your existing project and adapt paths / config.yaml as needed.
"""

import os
import json
import logging
import tempfile
from pathlib import Path
from datetime import datetime
import pytz
import pandas as pd
from openpyxl import load_workbook
import yaml
from bet_tracker_updater import update_master_from_dataframes

# Import existing helper functions from your current module if available.
# For portability we duplicate small helpers here; if you re-use your module, import instead.

PERTH = pytz.timezone("Australia/Perth")

# -------------------------------------------------------------
# Configurable paths & constants
# -------------------------------------------------------------

from constants import PROJECT_ROOT
EXPORT_ROOT = PROJECT_ROOT / "data" / "exports"
READY_ROOT = PROJECT_ROOT / "data" / "ready"
CACHE_PATH = PROJECT_ROOT / "processed_cache.json"
CONFIG_PATH = PROJECT_ROOT / "config.yaml"
FORMULA_PATH = PROJECT_ROOT / "formulas.yaml" 
date = datetime.now().strftime("%Y-%m-%d")
log_path = PROJECT_ROOT / "logs" / f"bookie_grabber_{date}.log"

# Ensure directories exist
for p in (EXPORT_ROOT, READY_ROOT, log_path.parent):
    p.mkdir(parents=True, exist_ok=True)

# Setup logging
logging.basicConfig(
    filename=str(log_path),
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger("bookie_grabber")

# -------------------------------------------------------------
# Utilities
# -------------------------------------------------------------

def load_yaml_safe(path: Path):
    if not path.exists():
        return {}
    try:
        with path.open("r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    except Exception as e:
        logger.error(f"Failed to load YAML {path}: {e}")
        return {}


def load_json_safe(path: Path):
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        logger.warning(f"Failed to read JSON {path}: {e}")
        return {}


def atomic_write(path: Path, data: bytes):
    """Write a file atomically."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(dir=str(path.parent), delete=False) as tf:
        tf.write(data)
        tmp = Path(tf.name)
    tmp.replace(path)


def iso_now_perth():
    return datetime.now(PERTH).isoformat()


# -------------------------------------------------------------
# Cache (processed events)
# -------------------------------------------------------------

def load_processed_cache(path: Path = CACHE_PATH):
    d = load_json_safe(path)
    # ensure keys are strings
    return {str(k): v for k, v in d.items()}


def save_processed_cache(cache: dict, path: Path = CACHE_PATH):
    try:
        atomic_write(path, json.dumps(cache, indent=2, ensure_ascii=False).encode("utf-8"))
    except Exception as e:
        logger.error(f"Could not save cache to {path}: {e}")


def mark_events_processed(event_ids, cache_path: Path = CACHE_PATH):
    cache = load_processed_cache(cache_path)
    ts = iso_now_perth()
    for eid in event_ids:
        cache[str(eid)] = ts
    save_processed_cache(cache, cache_path)


# -------------------------------------------------------------
# Excel export helpers
# -------------------------------------------------------------

def make_datetimes_timezone_naive(df: pd.DataFrame, datetime_cols: list):
    """Convert tz-aware datetime columns to naive (drop tzinfo) for Excel."""
    df_copy = df.copy()
    for col in datetime_cols:
        if col in df_copy.columns:
            df_copy[col] = pd.to_datetime(df_copy[col], errors='coerce').dt.tz_localize(None)
    return df_copy

def write_ready_workbook(date: datetime, league_slug:str,  totals_df: pd.DataFrame, btts_df: pd.DataFrame, out_dir: Path = READY_ROOT):
    """Writes a single workbook for the given date with multiple sheets."""
    DROP_COLS = ["event_id", "bf_team_name_home", "bf_team_name_away", "bf_merge_key", "marketid", "event"]
    date_str = date.strftime("%Y-%m-%d")
    out_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(PERTH).strftime("%H%M%S")
    filename = out_dir / f"ready_games_{league_slug}_{date_str}_{timestamp}.xlsx"

    # odds columns to ensure are floats
    totals_odds_cols = ['Bet365_over_odds', 'Bet365_under_odds', 'Betfair_Exchange_over_odds', 'Betfair_Exchange_under_odds']
    btts_odds_cols = ['Bet365_no_odds', 'Bet365_yes_odds', 'Betfair_Exchange_no_odds', 'Betfair_Exchange_yes_odds']

    # make sure the odds columns are float
    for col in totals_odds_cols:
        if col in totals_df.columns:
            totals_df[col] = pd.to_numeric(totals_df[col], errors='coerce')
    for col in btts_odds_cols:
        if col in btts_df.columns:
            btts_df[col] = pd.to_numeric(btts_df[col], errors='coerce')

    try:
        with pd.ExcelWriter(filename, engine="openpyxl", datetime_format="YYYY-MM-DD HH:MM:SS") as writer:
            
            datetime_cols = ["match_time", "odds_time"]  # adjust to actual datetime columns

            if not totals_df.empty:
                totals_df_naive = make_datetimes_timezone_naive(totals_df, datetime_cols)
                totals_df_naive.to_excel(writer, sheet_name="totals_ready", index=False)
            else:
                pd.DataFrame(columns=["event_id"]).to_excel(writer, sheet_name="totals_ready", index=False)

            if not btts_df.empty:
                btts_df_naive = make_datetimes_timezone_naive(btts_df, datetime_cols)
                btts_df_naive.to_excel(writer, sheet_name="btts_ready", index=False)
            else:
                pd.DataFrame(columns=["event_id"]).to_excel(writer, sheet_name="btts_ready", index=False)

            meta = {
                "generated_at": iso_now_perth(),
                "totals_rows": int(len(totals_df)),
                "btts_rows": int(len(btts_df)),
            }
            pd.DataFrame([meta]).to_excel(writer, sheet_name="meta", index=False)

        logger.info(f"Wrote ready workbook: {filename}")
        return filename
    except Exception as e:
        logger.exception(f"Failed to write ready workbook: {e}")
        raise


# -------------------------------------------------------------
# Master CSV archival
# -------------------------------------------------------------

def save_master_csv(df: pd.DataFrame, out_folder: Path, prefix: str, overwrite: bool = True):
    out_folder.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now(PERTH).strftime("%Y%m%d")
    out_path = out_folder / f"{prefix}_{date_str}.csv"

    # Skip writing if overwrite=False and file already exists
    if not overwrite and out_path.exists():
        logger.info(f"Master CSV already exists, skipping: {out_path}")
        return out_path

    try:
        df.to_csv(out_path, index=False)
        logger.info(f"Saved master CSV: {out_path}")
        return out_path
    except Exception as e:
        logger.exception(f"Failed to save master CSV to {out_path}: {e}")
        raise


# -------------------------------------------------------------
# High-level orchestration (plugs into your existing functions)
# -------------------------------------------------------------

def find_ready_games_from_master(master_df: pd.DataFrame, target_hours: int, processed_cache: dict):
    """Return rows from master_df which are ready and not yet processed.

    Assumes master_df has 'event_id', 'match_time', 'odds_time' columns and tz-aware datetimes.
    """
    if master_df.empty:
        return pd.DataFrame()

    # ready if hours_until_KO <= target_hours and not already processed
    mask_ready = master_df["hours_until_KO"] <= float(target_hours)
    mask_not_processed = ~master_df["event_id"].astype(str).isin(set(processed_cache.keys()))

    ready = master_df[mask_ready & mask_not_processed].copy()
    return ready


def expand_formula(template: str, row: int, bf_col: str | None = None, rpd_col: str | None = None) -> str:
    """
    Expands an Excel formula template for a given row:
        - replaces cell refs (L2/R2/H2/T2/I2/J2) with the correct row
        - replaces {ODDS} 
        - applies any custom substitutions
    """
    # Replace generic cell refs
    formula = (
        template
        .replace("L2", f"L{row}")
        .replace("R2", f"R{row}")
        .replace("H2", f"H{row}")
        .replace("T2", f"T{row}")
        .replace("I2", f"I{row}")
        .replace("J2", f"J{row}")
        .replace("K2", f"K{row}")
    )
    # Replace {ODDS}
    if bf_col and rpd_col:
        formula = formula.replace("{ODDS}", bf_col).replace("{RPD}", rpd_col)

    return formula

def apply_excel_formulas(workbook_path: Path, config: dict):
    wb = load_workbook(workbook_path)
    # -------- TOTALS --------
    if "totals_ready" in wb.sheetnames:
        ws = wb["totals_ready"]
        ws["T1"] = "o/u"
        ws["U1"] = "stake"
        for row in range(2, ws.max_row + 1):
            # O/U formula
            ws[f"T{row}"] = expand_formula(config["excel_formulas"]["O/U"], row)
            try:
                over_rpd = float(ws[f"K{row}"].value)
                under_rpd = float(ws[f"L{row}"].value)
            except (TypeError, ValueError):
                continue
            if over_rpd < under_rpd:
                bf_col = f"I{row}"
                rpd_col = f"K{row}"
            else:
                bf_col = f"J{row}"
                rpd_col = f"L{row}"

            line = ws[f"Q{row}"].value
            if line == "Over/Under 1.5 Goals":
                ws[f"U{row}"] = expand_formula(config["excel_formulas"]["totals"]["g1_5"], row, bf_col, rpd_col)
            elif line == "Over/Under 2.5 Goals":
                ws[f"U{row}"] = expand_formula(config["excel_formulas"]["totals"]["g2_5"], row, bf_col, rpd_col)
            else:
                ws[f"U{row}"] = expand_formula(config["excel_formulas"]["totals"]["g3_5"], row, bf_col, rpd_col)
    # -------- BTTS --------
    if "btts_ready" in wb.sheetnames:
        ws = wb["btts_ready"]
        ws["T1"] = "o/u"
        ws["U1"] = "stake"
        for row in range(2, ws.max_row + 1):
            # O/U formula
            ws[f"T{row}"] = expand_formula(config["excel_formulas"]["O/U"], row)
            over_rpd = ws[f"K{row}"].value
            under_rpd = ws[f"L{row}"].value
            # bf_col = f"J{row}" if over_rpd < under_rpd else f"I{row}"
            if over_rpd < under_rpd:
                bf_col = f"J{row}"
                rpd_col = f"K{row}"
            else:
                bf_col = f"I{row}"
                rpd_col = f"L{row}"
            # BTTS stake formula
            ws[f"U{row}"] = expand_formula(config["excel_formulas"]["btts"]["stake"], row, bf_col, rpd_col)

    wb.save(workbook_path)


# -------------------------------------------------------------
# Entrypoint glue to your existing pipeline
# -------------------------------------------------------------
# The expectation is you will call your existing process_league() (or adapt) so that
# it writes the master totals and btts CSVs (per-league/day) and returns DataFrames
# for the current snapshot. We'll provide a wrapper that takes those DataFrames,
# identifies "ready" matches, writes the Excel workbook and marks them as processed.
# -------------------------------------------------------------

def run_postprocessing_and_exports(league_slug: str, totals_master: pd.DataFrame, btts_master: pd.DataFrame, target_hours: int = 9):
    """
    Given the fresh master tables for a league, identify ready matches and export a workbook.

    - Totals and BTTS are never blindly merged; columns are preserved.
    - merged_ready sheet is a reference concat for inspection.
    - Marks ready events as processed.

    Returns:
        Path to Excel workbook, or None if no ready games.
    """
    cache = load_processed_cache(CACHE_PATH)

    # Save full master CSVs for archive
    master_folder = EXPORT_ROOT / league_slug
    try:
        if not totals_master.empty:
            save_master_csv(totals_master, master_folder / "totals", "totals", overwrite=True)
        if not btts_master.empty:
            save_master_csv(btts_master, master_folder / "btts", "btts", overwrite=False)
    except Exception:
        logger.exception("Failed saving master CSVs")

    # Find ready games
    ready_totals = find_ready_games_from_master(totals_master, target_hours, cache) if not totals_master.empty else pd.DataFrame()
    ready_btts = find_ready_games_from_master(btts_master, target_hours, cache) if not btts_master.empty else pd.DataFrame()

    # Collect all ready event_ids
    ready_event_ids = set()
    if not ready_totals.empty:
        ready_event_ids.update(map(str, ready_totals["event_id"].tolist()))
    if not ready_btts.empty:
        ready_event_ids.update(map(str, ready_btts["event_id"].tolist()))

    if len(ready_event_ids) == 0:
        logger.info("No ready games this run")
        return None

    # Export Excel workbook
    workbook_path = write_ready_workbook(datetime.now(PERTH), league_slug, ready_totals, ready_btts)

    # check if workbook has required columns before applying formulas
    wb = load_workbook(workbook_path, read_only=True)
    wst = wb["totals_ready"]
    totals_headers = [cell.value for cell in wst[1]]
    wsb = wb["btts_ready"]
    btts_headers = [cell.value for cell in wsb[1]]
    totals_required_columns = {
        "Bet365_over_odds",
        "Bet365_under_odds",
        "Betfair_Exchange_over_odds",
        "Betfair_Exchange_under_odds",
    }
    btts_required_columns = {
        "Bet365_yes_odds",
        "Bet365_no_odds",
        "Betfair_Exchange_yes_odds",
        "Betfair_Exchange_no_odds",
    }
    all_columsns_present = True
    if (not totals_required_columns.issubset(set(totals_headers))) or (not btts_required_columns.issubset(set(btts_headers))):
        logger.error("Workbook missing required columns — skipping formula application")
        all_columsns_present = False
    
    if all_columsns_present:
        try:
            formulas = load_yaml_safe(FORMULA_PATH)
            apply_excel_formulas(workbook_path, formulas)
        except Exception:
            logger.exception("Failed applying Excel formulas")  

    # Update master bet tracker spreadsheet
    try:
        appended = update_master_from_dataframes(ready_totals, ready_btts)
        if appended > 0:
            logger.info(f"Master bet tracker updated: {appended} rows appended")
    except Exception:
        logger.exception("Failed updating master bet tracker")

    # Mark processed events
    mark_events_processed(list(ready_event_ids), CACHE_PATH)

    return workbook_path