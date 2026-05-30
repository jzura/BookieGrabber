"""Hypothetical: bet >3.5G (Over) and BTTS Yes using the SAME core qualifying rules
currently used for the Under/No contrarian strategy (vol bounds, BF>1.45, RPD tiers).

Window: April + May 2026.
"""
import sys
from datetime import date, datetime
from openpyxl import load_workbook

sys.path.insert(0, "/Users/Joel/REPOS/BookieGrabber")
from strategy_config import (
    VOL_MIN, VOL_MAX, BF_MIN, BF_TIER1, RPD_TIER1, RPD_TIER2,
    compute_rpd, BF_SM_DISCOUNT_TIERS, estimate_sm_odds,
)

MASTER = "/Users/Joel/Desktop/EFB_Master_Bet_Tracker_VS Code.xlsx"
START, END = date(2026, 4, 1), date(2026, 5, 31)


def qualifies(bet_type, bf, vol, rpd):
    """Same core qualification used for Under/No strategy."""
    if bet_type not in ("3.5G", "BTTS"):
        return False
    if vol is None or not (VOL_MIN <= vol <= VOL_MAX):
        return False
    if bf is None or bf <= BF_MIN:
        return False
    if rpd is None:
        return False
    if bf <= BF_TIER1 and rpd > RPD_TIER1:
        return False
    if bf > BF_TIER1 and rpd > RPD_TIER2:
        return False
    return True


def main():
    wb = load_workbook(MASTER, read_only=True, data_only=True)
    ws = wb.active

    counted = 0
    wins = 0
    losses = 0
    pushes = 0
    total_pl = 0.0   # in units (€1 stake = 1 unit)
    by_type = {"3.5G": {"n": 0, "w": 0, "pl": 0.0},
               "BTTS": {"n": 0, "w": 0, "pl": 0.0}}

    rows_inspected = 0
    rows_no_result = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        bt = row[0]
        dt = row[1]
        pred = row[7]
        o365 = row[8]
        bf = row[9]
        vol = row[10]
        ng = row[13]  # home goals
        ag = row[14]  # away goals

        if bt not in ("3.5G", "BTTS"):
            continue
        if not isinstance(dt, datetime):
            continue
        d = dt.date()
        if not (START <= d <= END):
            continue
        # Mirror strategy: prediction=1 means market favours Over/Yes;
        # apply same RPD tier qualification.
        if pred != 1:
            continue
        rpd = compute_rpd(o365, bf)
        if not qualifies(bt, bf, vol, rpd):
            continue

        # Need results
        if ng is None or ag is None or ng == "" or ag == "":
            rows_no_result += 1
            continue
        try:
            ng = int(ng); ag = int(ag)
        except (TypeError, ValueError):
            rows_no_result += 1
            continue

        # Compute Over/Yes outcome
        if bt == "3.5G":
            won = (ng + ag) >= 4
        else:  # BTTS Yes
            won = (ng > 0 and ag > 0)

        # Use SM-discounted fill price (same approach the live strategy assumes)
        fill_odds = estimate_sm_odds(bf)
        if won:
            pl = (fill_odds - 1.0)
            wins += 1
            by_type[bt]["w"] += 1
        else:
            pl = -1.0
            losses += 1
        total_pl += pl
        by_type[bt]["n"] += 1
        by_type[bt]["pl"] += pl
        counted += 1

    wb.close()

    print(f"Window: {START} to {END}")
    print(f"Qualifying settled bets: {counted}")
    print(f"  Wins:   {wins}")
    print(f"  Losses: {losses}")
    print(f"  Skipped (no result):  {rows_no_result}")
    print()
    print(f"P/L per 1-unit stake (SM-discounted fills, no haircut beyond tier discount):")
    print(f"  TOTAL: {total_pl:+.3f} units")
    print(f"  ROI:   {(total_pl/counted*100 if counted else 0):+.2f}%")
    print()
    print("By market:")
    for bt, s in by_type.items():
        n = s["n"]; w = s["w"]; pl = s["pl"]
        wr = w / n * 100 if n else 0
        roi = pl / n * 100 if n else 0
        print(f"  {bt:5} n={n:4}  WR={wr:5.1f}%  P/L={pl:+8.2f}u  ROI={roi:+5.2f}%")
    print()
    print(f"At €250/unit (weekend rate): €{total_pl*250:+,.0f}")
    print(f"At €500/unit (midweek rate): €{total_pl*500:+,.0f}")
    # Rough blend assuming ~30% bets on weekend, 70% midweek (just to give shape)
    blend = total_pl * (0.3*250 + 0.7*500)
    print(f"At blended rate (30% wknd / 70% midweek): €{blend:+,.0f}")


if __name__ == "__main__":
    main()
