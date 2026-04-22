"""
Backtest engine for the bet tracker strategy.

Reads historical bet data from the master spreadsheet, applies a given
parameter configuration, and computes P&L statistics.

This is a pure simulation — it doesn't modify the master spreadsheet.
"""

import os
from pathlib import Path
from datetime import datetime, date, timedelta
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Optional

import openpyxl

MASTER_PATH = Path.home() / "Desktop" / "EFB_Master_Bet_Tracker_VS Code.xlsx"
MASTER_SHEET = "Master Bet Tracker"


# -------------------------------------------------------------
# Parameter configuration
# -------------------------------------------------------------

@dataclass
class StrategyParams:
    """Tunable parameters for backtesting/optimization.
    Defaults here are the OPTIMIZER search defaults — they intentionally differ
    from production values in strategy_config.py. The optimizer explores wider
    ranges to find optimal settings."""
    # Volume range
    vol_min: float = 40.0
    vol_max: float = 1100.0

    # BF lower bound
    bf_min: float = 1.30

    # BF tier breakpoints
    bf_tier1: float = 1.80   # boundary between low and mid
    bf_tier2: float = 2.50   # boundary between mid and high

    # RPD thresholds for each tier
    rpd_low: float = 1.0     # for BF <= bf_tier1
    rpd_mid: float = 2.0     # for bf_tier1 < BF <= bf_tier2
    rpd_high: float = 3.5    # for BF > bf_tier2

    # Fade thresholds (per-market)
    btts_fade_rpd: float = 5.0
    g15_fade_rpd: float = 4.6

    # Double-stake settings
    double_stake_rpd: float = 1.0       # threshold to qualify for 2x
    double_stake_min_count: int = 2     # min core qualifying bets per match

    def to_dict(self):
        return {
            "vol_min": self.vol_min,
            "vol_max": self.vol_max,
            "bf_min": self.bf_min,
            "bf_tier1": self.bf_tier1,
            "bf_tier2": self.bf_tier2,
            "rpd_low": self.rpd_low,
            "rpd_mid": self.rpd_mid,
            "rpd_high": self.rpd_high,
            "btts_fade_rpd": self.btts_fade_rpd,
            "g15_fade_rpd": self.g15_fade_rpd,
            "double_stake_rpd": self.double_stake_rpd,
            "double_stake_min_count": self.double_stake_min_count,
        }


@dataclass
class BacktestStats:
    """P&L stats for a backtest run."""
    n_bets: int = 0
    total_staked: float = 0.0
    total_returned: float = 0.0
    total_profit: float = 0.0
    n_wins: int = 0
    n_losses: int = 0
    n_double_stakes: int = 0
    by_market: dict = field(default_factory=dict)

    @property
    def roi(self) -> float:
        return (self.total_profit / self.total_staked * 100) if self.total_staked > 0 else 0.0

    @property
    def win_rate(self) -> float:
        return (self.n_wins / self.n_bets * 100) if self.n_bets > 0 else 0.0

    def to_dict(self):
        return {
            "n_bets": self.n_bets,
            "total_staked": round(self.total_staked, 2),
            "total_returned": round(self.total_returned, 2),
            "total_profit": round(self.total_profit, 2),
            "roi_pct": round(self.roi, 2),
            "win_rate_pct": round(self.win_rate, 2),
            "n_wins": self.n_wins,
            "n_losses": self.n_losses,
            "n_double_stakes": self.n_double_stakes,
        }


# -------------------------------------------------------------
# Data loading
# -------------------------------------------------------------

@dataclass
class BetRow:
    """A historical bet row with all the data needed to backtest."""
    bet_type: str        # "1.5G" / "2.5G" / "3.5G" / "BTTS"
    date: date
    home: str
    away: str
    prediction: int      # 0 or 1
    odds_365: float
    bf: float
    volume: float
    rpd: float
    goals: Optional[int] = None     # M column — total goals
    result: Optional[int] = None    # N column — 0 or 1 (BTTS direct, or computed for goals markets)


def _compute_rpd(odds_365, bf):
    try:
        i_val, j_val = float(odds_365), float(bf)
        if i_val > j_val:
            return 1.0
        pct = abs(i_val - j_val) / ((i_val + j_val) / 2) * 100
        return 1.0 if pct < 1 else round(pct, 3)
    except (ValueError, TypeError, ZeroDivisionError):
        return None


def _compute_goals_result(bet_type: str, prediction: int, total_goals: int) -> int:
    """For a goals market, compute whether the bet won (1) or lost (0)."""
    line = float(bet_type.replace("G", ""))   # 1.5, 2.5, 3.5
    if prediction == 1:  # Over
        return 1 if total_goals > line else 0
    else:  # Under
        return 1 if total_goals < line else 0


def load_historical_bets(master_path: Path = MASTER_PATH) -> list[BetRow]:
    """Read all bet rows from the master spreadsheet."""
    wb = openpyxl.load_workbook(master_path, data_only=True)
    ws = wb[MASTER_SHEET]

    bets = []
    for r in range(2, ws.max_row + 1):
        bt = ws.cell(row=r, column=1).value
        if bt is None:
            break

        d = ws.cell(row=r, column=2).value
        if isinstance(d, datetime):
            d = d.date()
        if not isinstance(d, date):
            continue

        home = ws.cell(row=r, column=3).value
        away = ws.cell(row=r, column=4).value
        pred = ws.cell(row=r, column=8).value
        odds_365 = ws.cell(row=r, column=9).value
        bf = ws.cell(row=r, column=10).value
        vol = ws.cell(row=r, column=11).value
        goals = ws.cell(row=r, column=13).value
        result = ws.cell(row=r, column=14).value

        try:
            odds_365_f = float(odds_365)
            bf_f = float(bf)
            vol_f = float(vol) if vol is not None else 0.0
        except (ValueError, TypeError):
            continue

        if pred not in (0, 1):
            continue

        rpd = _compute_rpd(odds_365_f, bf_f)
        if rpd is None:
            continue

        # Parse goals if present
        try:
            goals_int = int(goals) if goals not in (None, "") else None
        except (ValueError, TypeError):
            goals_int = None

        # Parse result — for goals markets it's a formula; for BTTS it's a direct value
        try:
            result_int = int(result) if isinstance(result, (int, float)) else None
        except (ValueError, TypeError):
            result_int = None

        # If we have goals but no result for goals markets, compute it
        if bt in ("1.5G", "2.5G", "3.5G") and result_int is None and goals_int is not None:
            result_int = _compute_goals_result(bt, pred, goals_int)

        bets.append(BetRow(
            bet_type=bt,
            date=d,
            home=str(home) if home else "",
            away=str(away) if away else "",
            prediction=int(pred),
            odds_365=odds_365_f,
            bf=bf_f,
            volume=vol_f,
            rpd=rpd,
            goals=goals_int,
            result=result_int,
        ))

    wb.close()
    return bets


# -------------------------------------------------------------
# Strategy logic (matching master sheet formulas)
# -------------------------------------------------------------

def is_core_qualifying(bet: BetRow, params: StrategyParams) -> bool:
    """Core contrarian: 1.5G/3.5G/BTTS, pred=0, vol/BF/RPD schedule."""
    if bet.bet_type not in ("1.5G", "3.5G", "BTTS"):
        return False
    if bet.prediction != 0:
        return False
    if not (params.vol_min <= bet.volume <= params.vol_max):
        return False
    if bet.bf <= params.bf_min:
        return False

    if bet.bf <= params.bf_tier1:
        return bet.rpd <= params.rpd_low
    elif bet.bf <= params.bf_tier2:
        return bet.rpd <= params.rpd_mid
    else:
        return bet.rpd <= params.rpd_high


def is_btts_fade(bet: BetRow, params: StrategyParams) -> bool:
    """BTTS fade: prediction=0 (No predicted), RPD >= threshold. Bets Yes (fading No)."""
    if bet.bet_type != "BTTS":
        return False
    if bet.prediction != 0:
        return False
    if bet.rpd < params.btts_fade_rpd:
        return False
    if not (params.vol_min <= bet.volume <= params.vol_max):
        return False
    return True


def is_15g_fade(bet: BetRow, params: StrategyParams) -> bool:
    """1.5G fade: prediction=1 (Over) with RPD >= threshold. Bets Under."""
    if bet.bet_type != "1.5G":
        return False
    if bet.prediction != 1:
        return False
    if bet.rpd < params.g15_fade_rpd:
        return False
    if not (params.vol_min <= bet.volume <= params.vol_max):
        return False
    return True


def detect_btts_conflicts(bets: list[BetRow]) -> set[int]:
    """Return indices of BTTS rows that should be skipped due to conflict
    (all 3 goals markets predict Over)."""
    pred_lookup = {}
    for i, b in enumerate(bets):
        pred_lookup[(b.date, b.home, b.away, b.bet_type)] = b.prediction

    conflicts = set()
    for i, b in enumerate(bets):
        if b.bet_type == "BTTS" and b.prediction == 0:
            if (pred_lookup.get((b.date, b.home, b.away, "1.5G")) == 1 and
                pred_lookup.get((b.date, b.home, b.away, "2.5G")) == 1 and
                pred_lookup.get((b.date, b.home, b.away, "3.5G")) == 1):
                conflicts.add(i)
    return conflicts


# -------------------------------------------------------------
# Return calculation (matching master sheet Return formula)
# -------------------------------------------------------------

# Haircut on theoretical opposite odds for fade bets.
# Based on 4 actual SM fade fills: real odds average 3.6% below theoretical.
# Using 4% for safety (imported from strategy_config).
from strategy_config import FADE_ODDS_HAIRCUT as FADE_OPPOSITE_ODDS_HAIRCUT


def _commission(odds: float) -> float:
    """SportsMarket tiered commission."""
    if odds <= 1.5:
        return 0.01
    if odds <= 2.8:
        return 0.02
    if odds <= 3.5:
        return 0.03
    return 0.04


def calculate_return(bet: BetRow, stake: float, is_fade: bool) -> float:
    """
    Calculate the return for a bet given a stake.
    For core bets this matches the spreadsheet's Return formula exactly.
    For fade bets we apply a conservative 3% haircut to the theoretical
    opposite odds to account for bookmaker margin (which the spreadsheet
    formula doesn't account for, since the spreadsheet uses actual SM
    odds entered after placing the bet).
    """
    if bet.result is None:
        return 0.0  # unsettled

    if is_fade:
        # Fade bets win when result = 0 (i.e. the prediction was wrong)
        if bet.result == 0:
            opp_odds = (1 / (1 - 1 / bet.bf)) * (1 - FADE_OPPOSITE_ODDS_HAIRCUT)
            cm = _commission(opp_odds)
            return stake * (1 + (opp_odds - 1) * (1 - cm))
        else:
            return 0.0
    else:
        # Core: wins when result = 1 (prediction correct)
        if bet.result == 1:
            cm = _commission(bet.bf)
            return stake * (1 + (bet.bf - 1) * (1 - cm))
        else:
            return 0.0


# -------------------------------------------------------------
# Backtest runner
# -------------------------------------------------------------

def run_backtest(bets: list[BetRow], params: StrategyParams,
                 date_from: Optional[date] = None,
                 date_to: Optional[date] = None) -> BacktestStats:
    """
    Run a backtest over the given bets with the given parameters.
    Optionally filter by date range.
    """
    # Filter by date
    in_range = [b for b in bets
                if (date_from is None or b.date >= date_from)
                and (date_to is None or b.date <= date_to)]

    # Detect BTTS conflicts (still needs full row context for the same match)
    conflicts = detect_btts_conflicts(in_range)

    # Identify core qualifying bets
    core_indices = set()
    for i, b in enumerate(in_range):
        if i in conflicts:
            continue
        if is_core_qualifying(b, params):
            core_indices.add(i)

    # Compute double-stake eligibility (per-match confluence)
    match_core_count = defaultdict(int)
    for i in core_indices:
        b = in_range[i]
        match_core_count[(b.date, b.home, b.away)] += 1

    double_stake = set()
    for i in core_indices:
        b = in_range[i]
        if (b.rpd <= params.double_stake_rpd
                and match_core_count[(b.date, b.home, b.away)] >= params.double_stake_min_count):
            double_stake.add(i)

    # Under 2.5G piggyback: matches where 1.5G Under qualifies as core → flag 2.5G Under
    matches_with_core_15g = set()
    for i in core_indices:
        b = in_range[i]
        if b.bet_type == "1.5G" and b.prediction == 0:
            matches_with_core_15g.add((b.date, b.home, b.away))

    piggyback_25g = set()
    for i, b in enumerate(in_range):
        if b.bet_type == "2.5G" and b.prediction == 0 and b.bf > params.bf_min:
            if (b.date, b.home, b.away) in matches_with_core_15g:
                piggyback_25g.add(i)

    # Simulate P&L
    stats = BacktestStats()
    by_market = defaultdict(lambda: BacktestStats())

    for i, b in enumerate(in_range):
        if i in conflicts:
            continue
        if b.result is None:
            continue  # unsettled — skip

        # Determine if this bet qualifies and what the stake is
        stake = 0.0
        is_fade = False

        if i in core_indices:
            stake = 2.0 if i in double_stake else 1.0
        elif i in piggyback_25g:
            stake = 1.0  # Under 2.5G piggyback — always 1x, same return calc as core
        elif is_btts_fade(b, params):
            stake = 1.0
            is_fade = True
        elif is_15g_fade(b, params):
            stake = 1.0
            is_fade = True

        if stake == 0:
            continue

        ret = calculate_return(b, stake, is_fade)
        profit = ret - stake

        stats.n_bets += 1
        stats.total_staked += stake
        stats.total_returned += ret
        stats.total_profit += profit
        if ret > 0:
            stats.n_wins += 1
        else:
            stats.n_losses += 1
        if stake == 2.0:
            stats.n_double_stakes += 1

        # Per-market breakdown
        m = by_market[b.bet_type]
        m.n_bets += 1
        m.total_staked += stake
        m.total_returned += ret
        m.total_profit += profit
        if ret > 0:
            m.n_wins += 1
        else:
            m.n_losses += 1
        if stake == 2.0:
            m.n_double_stakes += 1

    stats.by_market = {k: v.to_dict() for k, v in by_market.items()}
    return stats


# -------------------------------------------------------------
# CLI for sanity testing
# -------------------------------------------------------------

if __name__ == "__main__":
    print("Loading historical bets...")
    bets = load_historical_bets()
    print(f"Loaded {len(bets)} bet rows")

    settled = [b for b in bets if b.result is not None]
    print(f"Settled rows (with result): {len(settled)}")

    print("\nRunning backtest with current default parameters...")
    params = StrategyParams()
    stats = run_backtest(bets, params)

    print(f"\n=== Overall ===")
    for k, v in stats.to_dict().items():
        print(f"  {k}: {v}")

    print(f"\n=== By Market ===")
    for market, m_stats in stats.by_market.items():
        print(f"  {market}: bets={m_stats['n_bets']}, profit=${m_stats['total_profit']}, ROI={m_stats['roi_pct']}%")