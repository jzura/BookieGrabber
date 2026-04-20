"""
Apply optimized parameters to the master spreadsheet.

Reads a config from a CSV report (the output of optimize.py) and rewrites
the stake formulas in column O of the master sheet to match.

Also writes the chosen config to optimizer/active_params.json so we have
a record of what's currently applied.

Usage:
    python optimizer/apply_params.py --config optimizer/results/optim_YYYYMMDD_HHMMSS.csv --rank 3
    python optimizer/apply_params.py --config <path> --rank baseline    # restore defaults
    python optimizer/apply_params.py --dry-run --config <path> --rank 3 # preview only
"""

import argparse
import csv
import json
import sys
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from backtest import StrategyParams, load_historical_bets, run_backtest

MASTER_PATH = Path.home() / "Desktop" / "EFB_Master_Bet_Tracker_VS Code.xlsx"
MASTER_SHEET = "Master Bet Tracker"
ACTIVE_PARAMS_PATH = Path(__file__).resolve().parent / "active_params.json"


def load_config_from_csv(csv_path: Path, rank: str) -> StrategyParams:
    """Load a parameter config from a report CSV by rank."""
    with open(csv_path, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["rank"] == rank:
                params = StrategyParams()
                for k in params.to_dict().keys():
                    if k in row and row[k]:
                        val = row[k]
                        try:
                            if k == "double_stake_min_count":
                                setattr(params, k, int(float(val)))
                            else:
                                setattr(params, k, float(val))
                        except ValueError:
                            pass
                return params
    raise ValueError(f"Rank '{rank}' not found in {csv_path}")


# -------------------------------------------------------------
# Stake formula generators (parameterized version)
# -------------------------------------------------------------

def core_cond(r: int, p: StrategyParams) -> str:
    return (
        f'AND(OR(A{r}="1.5G",A{r}="3.5G",A{r}="BTTS"),H{r}=0,'
        f'K{r}>={p.vol_min},K{r}<={p.vol_max},J{r}>{p.bf_min},'
        f'OR('
        f'AND(J{r}<={p.bf_tier1},L{r}<={p.rpd_low}),'
        f'AND(J{r}>{p.bf_tier1},J{r}<={p.bf_tier2},L{r}<={p.rpd_mid}),'
        f'AND(J{r}>{p.bf_tier2},L{r}<={p.rpd_high})'
        f'))'
    )


def fade_cond(r: int, p: StrategyParams) -> str:
    return (
        f'AND(A{r}="BTTS",L{r}>={p.btts_fade_rpd},K{r}>={p.vol_min},K{r}<={p.vol_max}),'
        f'AND(A{r}="1.5G",H{r}=1,L{r}>={p.g15_fade_rpd},K{r}>={p.vol_min},K{r}<={p.vol_max})'
    )


def stake_formula(r: int, p: StrategyParams, is_conflict: bool, is_double_stake: bool,
                   is_25g_piggyback: bool = False) -> str:
    if is_conflict:
        return '=""'
    if is_25g_piggyback:
        return f'=IF(AND(A{r}="2.5G",H{r}=0,J{r}>{p.bf_min},P{r}<>""),1,"")'
    if is_double_stake:
        return f'=IF({core_cond(r, p)},2,IF(OR({fade_cond(r, p)}),1,""))'
    return f'=IF(OR({core_cond(r, p)},{fade_cond(r, p)}),1,"")'


# -------------------------------------------------------------
# Apply to master sheet
# -------------------------------------------------------------

def apply_to_sheet(params: StrategyParams, dry_run: bool = False):
    """Rewrite stake formulas in the master sheet using the given params."""
    print(f"\nLoading {MASTER_PATH}...")
    wb = openpyxl.load_workbook(MASTER_PATH)
    ws = wb[MASTER_SHEET]

    # Read all rows to compute conflict + double-stake sets in Python
    rows = {}
    for r in range(2, ws.max_row + 1):
        bt = ws.cell(row=r, column=1).value
        if bt is None:
            break
        d = ws.cell(row=r, column=2).value
        home = ws.cell(row=r, column=3).value
        away = ws.cell(row=r, column=4).value
        pred = ws.cell(row=r, column=8).value
        odds_365 = ws.cell(row=r, column=9).value
        bf = ws.cell(row=r, column=10).value
        vol = ws.cell(row=r, column=11).value
        if isinstance(d, datetime):
            d = d.date()
        try:
            i_val, j_val = float(odds_365), float(bf)
            if i_val > j_val:
                rpd = 1.0
            else:
                pct = abs(i_val - j_val) / ((i_val + j_val) / 2) * 100
                rpd = 1.0 if pct < 1 else round(pct, 3)
            vol_f = float(vol) if vol is not None else None
            bf_f = float(bf)
        except (ValueError, TypeError, ZeroDivisionError):
            rpd = None
            vol_f = None
            bf_f = None
        rows[r] = {"bt": bt, "date": d, "home": home, "away": away, "pred": pred,
                   "bf": bf_f, "vol": vol_f, "rpd": rpd}

    # BTTS conflicts (only — 3.5G filter is removed)
    pred_lookup = {(str(d["date"]), str(d["home"]), str(d["away"]), d["bt"]): d["pred"]
                   for d in rows.values()}
    conflicts = set()
    for r, d in rows.items():
        if d["bt"] == "BTTS" and d["pred"] == 0:
            k_d, k_h, k_a = str(d["date"]), str(d["home"]), str(d["away"])
            if (pred_lookup.get((k_d, k_h, k_a, "1.5G")) == 1 and
                pred_lookup.get((k_d, k_h, k_a, "2.5G")) == 1 and
                pred_lookup.get((k_d, k_h, k_a, "3.5G")) == 1):
                conflicts.add(r)

    # Core qualifying using the NEW params
    def is_core(d, r):
        if d["bt"] not in ("1.5G", "3.5G", "BTTS"): return False
        if d["pred"] != 0: return False
        if d["vol"] is None or not (params.vol_min <= d["vol"] <= params.vol_max): return False
        if d["bf"] is None or d["bf"] <= params.bf_min: return False
        if d["rpd"] is None: return False
        if d["bf"] <= params.bf_tier1 and d["rpd"] > params.rpd_low: return False
        if params.bf_tier1 < d["bf"] <= params.bf_tier2 and d["rpd"] > params.rpd_mid: return False
        if d["bf"] > params.bf_tier2 and d["rpd"] > params.rpd_high: return False
        if r in conflicts: return False
        return True

    core_rows = set()
    match_core_count = defaultdict(int)
    for r, d in rows.items():
        if is_core(d, r):
            core_rows.add(r)
            match_core_count[(str(d["date"]), str(d["home"]), str(d["away"]))] += 1

    double_stake = set()
    for r in core_rows:
        d = rows[r]
        mk = (str(d["date"]), str(d["home"]), str(d["away"]))
        if d["rpd"] == params.double_stake_rpd and match_core_count[mk] >= params.double_stake_min_count:
            double_stake.add(r)

    # Under 2.5G piggyback: matches where 1.5G Under qualifies as core
    matches_with_core_15g = set()
    for r in core_rows:
        d = rows[r]
        if d["bt"] == "1.5G" and d["pred"] == 0:
            matches_with_core_15g.add((str(d["date"]), str(d["home"]), str(d["away"])))

    piggyback_25g = set()
    for r, d in rows.items():
        if d["bt"] == "2.5G" and d["pred"] == 0:
            mk = (str(d["date"]), str(d["home"]), str(d["away"]))
            if mk in matches_with_core_15g and d["bf"] is not None and d["bf"] > params.bf_min:
                piggyback_25g.add(r)

    print(f"  Rows: {len(rows)}")
    print(f"  Core qualifying: {len(core_rows)}")
    print(f"  Double-stake: {len(double_stake)}")
    print(f"  BTTS conflicts: {len(conflicts)}")
    print(f"  2.5G piggyback: {len(piggyback_25g)}")

    if dry_run:
        print("\n[DRY RUN] No changes written.")
        wb.close()
        return

    # Write formulas
    print("\nRewriting stake formulas...")
    for r in rows:
        ws.cell(row=r, column=15, value=stake_formula(
            r, params, is_conflict=(r in conflicts), is_double_stake=(r in double_stake),
            is_25g_piggyback=(r in piggyback_25g),
        ))

    wb.save(MASTER_PATH)
    print(f"  Saved {len(rows)} stake formulas")

    # Save active params
    active = {
        "applied_at": datetime.now().isoformat(),
        "params": params.to_dict(),
    }
    ACTIVE_PARAMS_PATH.write_text(json.dumps(active, indent=2))
    print(f"  Active params recorded: {ACTIVE_PARAMS_PATH}")


# -------------------------------------------------------------
# CLI
# -------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Apply optimized parameters to master sheet")
    parser.add_argument("--config", type=str, required=True, help="Path to optimizer CSV report")
    parser.add_argument("--rank", type=str, required=True, help="Rank to apply (1, 2, ..., or 'baseline')")
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing")
    parser.add_argument("--yes", action="store_true", help="Skip confirmation prompt")
    args = parser.parse_args()

    csv_path = Path(args.config)
    if not csv_path.exists():
        print(f"ERROR: {csv_path} not found")
        return

    print(f"Loading config rank '{args.rank}' from {csv_path}...")
    params = load_config_from_csv(csv_path, args.rank)

    print("\nParameters to apply:")
    for k, v in params.to_dict().items():
        print(f"  {k:30s} {v}")

    # Sanity backtest on full data
    print("\nRunning backtest on full historical data...")
    bets = load_historical_bets()
    stats = run_backtest(bets, params)
    print(f"  All-time: {stats.n_bets} bets, ${stats.total_profit:.2f} profit, {stats.roi:.2f}% ROI")

    if args.dry_run:
        apply_to_sheet(params, dry_run=True)
        return

    if not args.yes:
        confirm = input("\nProceed and rewrite stake formulas in the master sheet? [y/N]: ")
        if confirm.strip().lower() != "y":
            print("Aborted.")
            return

    apply_to_sheet(params, dry_run=False)


if __name__ == "__main__":
    main()