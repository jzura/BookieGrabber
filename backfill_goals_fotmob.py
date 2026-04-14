"""
Fallback backfill: resolve rows still missing N/O using Fotmob's public
data API (returns every match globally for a given date, so it handles
UEFA qualifiers, Serbian/Czech/Croatian leagues, etc.).

Reads the master sheet, collects all unresolved fixtures directly
(doesn't rely on backfill_unresolved.csv — re-scans for anything still
blank), fetches Fotmob per unique date ±1, and fuzzy-matches team names.
"""

import csv
import logging
import sys
import time
from collections import defaultdict
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl
import pandas as pd
import requests

PROJECT_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(PROJECT_ROOT))

from results_updater import (
    MASTER_PATH,
    MASTER_SHEET,
    build_lookup,
    find_result,
)

LOG_DIR = PROJECT_ROOT / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)
LOG_PATH = LOG_DIR / f"backfill_fotmob_{datetime.now().strftime('%Y-%m-%d_%H%M')}.log"
UNRESOLVED_CSV = LOG_DIR / "backfill_unresolved_final.csv"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(str(LOG_PATH)), logging.StreamHandler()],
    force=True,
)
logger = logging.getLogger("backfill_fotmob")

FOTMOB_URL = "https://www.fotmob.com/api/data/matches?date={date}"
HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"}


def fetch_fotmob_day(d):
    """Return list of dicts {_date, _home, _away, _hg, _ag} for matches
    finished on day d (Fotmob uses UTC-ish date key)."""
    url = FOTMOB_URL.format(date=d.strftime("%Y%m%d"))
    try:
        r = requests.get(url, headers=HEADERS, timeout=20)
        if r.status_code != 200:
            logger.debug(f"Fotmob {d}: HTTP {r.status_code}")
            return []
        data = r.json()
    except Exception as e:
        logger.debug(f"Fotmob {d}: {e}")
        return []

    out = []
    for league in data.get("leagues", []):
        for m in league.get("matches", []):
            home = m.get("home") or {}
            away = m.get("away") or {}
            hs = home.get("score")
            as_ = away.get("score")
            if hs is None or as_ is None:
                continue
            try:
                hs = int(hs)
                as_ = int(as_)
            except (TypeError, ValueError):
                continue
            # Use the actual match date from status.utcTime if present
            status = m.get("status") or {}
            utc = status.get("utcTime")
            match_date = d
            if utc:
                try:
                    match_date = datetime.fromisoformat(utc.replace("Z", "+00:00")).date()
                except Exception:
                    pass
            out.append({
                "_date": match_date,
                "_home": home.get("longName") or home.get("name") or "",
                "_away": away.get("longName") or away.get("name") or "",
                "_hg": hs,
                "_ag": as_,
            })
    return out


def main():
    logger.info(f"Loading {MASTER_PATH}")
    wb = openpyxl.load_workbook(MASTER_PATH)
    ws = wb[MASTER_SHEET]

    fixtures = {}  # (date, home, away) -> {"comp": ..., "rows": [...]}
    for r in range(2, ws.max_row + 1):
        bt = ws.cell(row=r, column=1).value
        if bt is None:
            break
        n_val = ws.cell(row=r, column=14).value
        o_val = ws.cell(row=r, column=15).value
        if n_val not in (None, "") and o_val not in (None, ""):
            continue
        d = ws.cell(row=r, column=2).value
        home = ws.cell(row=r, column=3).value
        away = ws.cell(row=r, column=4).value
        comp = ws.cell(row=r, column=5).value
        if isinstance(d, datetime):
            d = d.date()
        if not isinstance(d, date) or d >= date.today():
            continue
        key = (d, home, away)
        fixtures.setdefault(key, {"comp": comp, "rows": []})["rows"].append(r)

    logger.info(f"Unresolved fixtures: {len(fixtures)}")
    if not fixtures:
        logger.info("Nothing to do")
        return

    # Collect unique dates ±1 to handle TZ / kickoff-after-midnight
    needed_dates = set()
    for (d, _h, _a) in fixtures.keys():
        needed_dates.add(d)
        needed_dates.add(d - timedelta(days=1))
        needed_dates.add(d + timedelta(days=1))

    logger.info(f"Fetching Fotmob for {len(needed_dates)} unique dates...")
    all_rows = []
    for i, d in enumerate(sorted(needed_dates), 1):
        rows = fetch_fotmob_day(d)
        all_rows.extend(rows)
        if i % 20 == 0:
            logger.info(f"  {i}/{len(needed_dates)} dates, {len(all_rows)} results so far")
        time.sleep(0.25)  # rate limit politely

    logger.info(f"Fotmob total results: {len(all_rows)}")
    if not all_rows:
        logger.error("No results from Fotmob — aborting")
        return

    df = pd.DataFrame(all_rows)
    lookup = build_lookup(df)

    resolved = 0
    rows_filled = 0
    unresolved = []

    for (d, home, away), meta in fixtures.items():
        hg, ag = find_result(lookup, d, home, away)
        if hg is None:
            unresolved.append({
                "date": d.isoformat(),
                "home": home,
                "away": away,
                "competition": meta["comp"],
                "rows": ";".join(str(r) for r in meta["rows"]),
            })
            continue
        for r in meta["rows"]:
            ws.cell(row=r, column=14, value=hg)
            ws.cell(row=r, column=15, value=ag)
            rows_filled += 1
        resolved += 1

    logger.info(f"Resolved: {resolved}/{len(fixtures)}")
    logger.info(f"Rows filled: {rows_filled}")
    logger.info(f"Still unresolved: {len(unresolved)}")

    if unresolved:
        with open(UNRESOLVED_CSV, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=["date", "home", "away", "competition", "rows"])
            writer.writeheader()
            writer.writerows(unresolved)
        logger.info(f"Final unresolved list: {UNRESOLVED_CSV}")
        # Also show breakdown by competition
        from collections import Counter
        by_comp = Counter(u["competition"] for u in unresolved)
        logger.info("Unresolved breakdown by competition:")
        for c, n in by_comp.most_common():
            logger.info(f"  {n:4d}  {c}")

    logger.info("Saving spreadsheet...")
    wb.save(MASTER_PATH)
    logger.info(f"Saved {MASTER_PATH}")


if __name__ == "__main__":
    main()