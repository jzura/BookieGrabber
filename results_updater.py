"""
Daily Results Updater

Fetches match results from football-data.co.uk and ESPN,
then populates column M (Goals) and column N (Result) in the
master bet tracker spreadsheet.

Runs daily at 07:00 via launchd.

Usage:
    python results_updater.py
"""

import os
import logging
import requests
import pandas as pd
from io import StringIO
from datetime import datetime, date, timedelta
from difflib import SequenceMatcher
from pathlib import Path
from collections import defaultdict

import openpyxl

# -------------------------------------------------------------
# Configuration
# -------------------------------------------------------------

MASTER_PATH = Path.home() / "Desktop" / "EFB_Master_Bet_Tracker_VS Code.xlsx"
MASTER_SHEET = "Master Bet Tracker"

PROJECT_ROOT = Path(__file__).resolve().parent
LOG_DIR = PROJECT_ROOT / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)
date_str = datetime.now().strftime("%Y-%m-%d")
LOG_PATH = LOG_DIR / f"results_updater_{date_str}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(str(LOG_PATH)),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger("results_updater")

# football-data.co.uk CSV sources
MAIN_CSVS = {
    "English Premier League": "https://www.football-data.co.uk/mmz4281/2526/E0.csv",
    "English Sky Bet Championship": "https://www.football-data.co.uk/mmz4281/2526/E1.csv",
    "French Ligue 1": "https://www.football-data.co.uk/mmz4281/2526/F1.csv",
    "German Bundesliga": "https://www.football-data.co.uk/mmz4281/2526/D1.csv",
    "Italian Serie A": "https://www.football-data.co.uk/mmz4281/2526/I1.csv",
    "Spanish La Liga": "https://www.football-data.co.uk/mmz4281/2526/SP1.csv",
    "Portuguese Primeira Liga": "https://www.football-data.co.uk/mmz4281/2526/P1.csv",
    "Belgian Pro League": "https://www.football-data.co.uk/mmz4281/2526/B1.csv",
    "Turkish Super League": "https://www.football-data.co.uk/mmz4281/2526/T1.csv",
    "Dutch Eredivisie": "https://www.football-data.co.uk/mmz4281/2526/N1.csv",
    "Greek Super League": "https://www.football-data.co.uk/mmz4281/2526/G1.csv",
    "Scottish Premiership": "https://www.football-data.co.uk/mmz4281/2526/SC0.csv",
}

NEW_CSVS = {
    "Austrian Bundesliga": "https://www.football-data.co.uk/new/AUT.csv",
    "Danish Superliga": "https://www.football-data.co.uk/new/DNK.csv",
    "Polish Ekstraklasa": "https://www.football-data.co.uk/new/POL.csv",
    "Romanian Liga I": "https://www.football-data.co.uk/new/ROU.csv",
    "Swiss Super League": "https://www.football-data.co.uk/new/SWZ.csv",
}

# ESPN API for leagues not on football-data.co.uk
ESPN_LEAGUES = {
    "UEFA Champions League": "uefa.champions",
    "UEFA Europa League": "uefa.europa",
    "UEFA Conference League": "uefa.europa.conf",
    "Croatian HNL": "cro.1",
    "Croatia - HNL": "cro.1",
    "Serbian SuperLiga": "srb.super_liga",
    "Serbia - Superliga": "srb.super_liga",
    "Serbian Super Liga": "srb.super_liga",
    "Czech First League": "cze.1",
    "Czechia - 1. Liga": "cze.1",
    "Norwegian Eliteserien": "nor.1",
    "Swedish Allsvenskan": "swe.1",
    "Swedish Cup": "swe.1",
    # Main leagues (ESPN fallback when football-data.co.uk CSVs are delayed)
    "English Premier League": "eng.1",
    "England - Championship": "eng.2",
    "English Championship League": "eng.2",
    "English Sky Bet Championship": "eng.2",
    "French Ligue 1": "fra.1",
    "France - Ligue 1": "fra.1",
    "German Bundesliga": "ger.1",
    "Germany - Bundesliga": "ger.1",
    "Italian Serie A": "ita.1",
    "Italy - Serie A": "ita.1",
    "Spanish La Liga": "esp.1",
    "Spain - LaLiga": "esp.1",
    "Portuguese Primeira Liga": "por.1",
    "Portugal - Liga Portugal": "por.1",
    "Belgian Pro League": "bel.1",
    "Belgium - Pro League": "bel.1",
    "Belgian First Division A": "bel.1",
    "Turkish Super League": "tur.1",
    "Turkish Super Lig": "tur.1",
    "Turkiye - Super Lig": "tur.1",
    "Dutch Eredivisie": "ned.1",
    "Netherlands - Eredivisie": "ned.1",
    "Netherlands Eredivisie": "ned.1",
    "Greek Super League": "gre.1",
    "Scottish Premiership": "sco.1",
    "Scotland - Premiership": "sco.1",
    "Austrian Bundesliga": "aut.1",
    "Austria - Bundesliga": "aut.1",
    "Danish Superliga": "den.1",
    "Denmark - Superliga": "den.1",
    "Polish Ekstraklasa": "pol.1",
    "Poland - Ekstraklasa": "pol.1",
    "Romanian Liga I": "rou.1",
    "Romania - Superliga": "rou.1",
    "Swiss Super League": "sui.1",
    "Switzerland - Super League": "sui.1",
}


# -------------------------------------------------------------
# Download results
# -------------------------------------------------------------

def download_football_data_csvs():
    """Download results from football-data.co.uk."""
    all_results = []

    for league, url in MAIN_CSVS.items():
        try:
            r = requests.get(url, timeout=15)
            r.raise_for_status()
            df = pd.read_csv(StringIO(r.text))
            results = pd.DataFrame({
                "_date": pd.to_datetime(df["Date"], dayfirst=True).dt.date,
                "_home": df["HomeTeam"],
                "_away": df["AwayTeam"],
                "_hg": pd.to_numeric(df["FTHG"], errors="coerce"),
                "_ag": pd.to_numeric(df["FTAG"], errors="coerce"),
            }).dropna(subset=["_hg", "_ag"])
            all_results.append(results)
            logger.info(f"football-data.co.uk {league}: {len(results)} matches")
        except Exception as e:
            logger.warning(f"football-data.co.uk {league}: {e}")

    for league, url in NEW_CSVS.items():
        try:
            r = requests.get(url, timeout=15)
            r.raise_for_status()
            df = pd.read_csv(StringIO(r.text))
            df = df[df["Season"] == "2025/2026"]
            results = pd.DataFrame({
                "_date": pd.to_datetime(df["Date"], dayfirst=True).dt.date,
                "_home": df["Home"],
                "_away": df["Away"],
                "_hg": pd.to_numeric(df["HG"], errors="coerce"),
                "_ag": pd.to_numeric(df["AG"], errors="coerce"),
            }).dropna(subset=["_hg", "_ag"])
            all_results.append(results)
            logger.info(f"football-data.co.uk {league}: {len(results)} matches")
        except Exception as e:
            logger.warning(f"football-data.co.uk {league}: {e}")

    return pd.concat(all_results, ignore_index=True) if all_results else pd.DataFrame()


def download_fotmob_results(dates):
    """Fetch all finished matches from Fotmob for the given set of dates.
    Fotmob returns every match globally for a date, so a handful of calls
    covers every league we track."""
    rows = []
    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"}
    for d in sorted(dates):
        url = f"https://www.fotmob.com/api/data/matches?date={d.strftime('%Y%m%d')}"
        try:
            r = requests.get(url, headers=headers, timeout=20)
            if r.status_code != 200:
                logger.warning(f"Fotmob {d}: HTTP {r.status_code}")
                continue
            data = r.json()
        except Exception as e:
            logger.warning(f"Fotmob {d}: {e}")
            continue

        for league in data.get("leagues", []):
            for m in league.get("matches", []):
                home = m.get("home") or {}
                away = m.get("away") or {}
                hs, as_ = home.get("score"), away.get("score")
                if hs is None or as_ is None:
                    continue
                try:
                    hs, as_ = int(hs), int(as_)
                except (TypeError, ValueError):
                    continue
                status = m.get("status") or {}
                utc = status.get("utcTime")
                match_date = d
                if utc:
                    try:
                        match_date = datetime.fromisoformat(utc.replace("Z", "+00:00")).date()
                    except Exception:
                        pass
                rows.append({
                    "_date": match_date,
                    "_home": home.get("longName") or home.get("name") or "",
                    "_away": away.get("longName") or away.get("name") or "",
                    "_hg": hs,
                    "_ag": as_,
                })

    logger.info(f"Fotmob: {len(rows)} results fetched across {len(dates)} dates")
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def download_espn_results(needed_dates):
    """
    Download results from ESPN API for leagues not covered by football-data.co.uk.
    needed_dates: dict of {espn_league_code: set of dates}
    """
    all_results = []

    for league_code, dates in needed_dates.items():
        for d in sorted(dates):
            date_str = d.strftime("%Y%m%d")
            url = f"https://site.api.espn.com/apis/site/v2/sports/soccer/{league_code}/scoreboard?dates={date_str}"
            try:
                r = requests.get(url, timeout=10)
                r.raise_for_status()
                data = r.json()
                for ev in data.get("events", []):
                    comps = ev.get("competitions", [])
                    if not comps:
                        continue
                    teams = comps[0].get("competitors", [])
                    if len(teams) != 2:
                        continue

                    home = away = None
                    for t in teams:
                        info = {"name": t["team"]["displayName"], "score": t.get("score", "")}
                        if t.get("homeAway") == "home":
                            home = info
                        else:
                            away = info

                    if home and away and home["score"] and away["score"]:
                        try:
                            all_results.append({
                                "_date": d,
                                "_home": home["name"],
                                "_away": away["name"],
                                "_hg": int(home["score"]),
                                "_ag": int(away["score"]),
                            })
                        except ValueError:
                            pass
            except Exception:
                pass

    logger.info(f"ESPN API: {len(all_results)} results fetched")
    return pd.DataFrame(all_results) if all_results else pd.DataFrame()


# -------------------------------------------------------------
# Matching logic
# -------------------------------------------------------------

def normalize(name):
    """Normalize team name for fuzzy matching."""
    if not isinstance(name, str):
        return ""
    return (name.lower()
        .replace("fc ", "").replace(" fc", "")
        .replace("sc ", "").replace(" sc", "")
        .replace("sv ", "").replace(" sv", "")
        .replace("fk ", "").replace(" fk", "")
        .replace("ac ", "").replace(" ac", "")
        .replace("1. ", "").replace(".", "").replace("'", "")
        .strip())


def build_lookup(results_df):
    """Build date-based lookup from results DataFrame."""
    lookup = {}
    for _, row in results_df.iterrows():
        d = row["_date"]
        if d not in lookup:
            lookup[d] = []
        lookup[d].append({
            "home": row["_home"],
            "away": row["_away"],
            "hg": int(row["_hg"]),
            "ag": int(row["_ag"]),
            "home_norm": normalize(row["_home"]),
            "away_norm": normalize(row["_away"]),
        })
    return lookup


def find_result(lookup, match_date, home_team, away_team):
    """Find match result, trying exact date and +/- 1 day."""
    home_n = normalize(home_team)
    away_n = normalize(away_team)

    for d_offset in [0, -1, 1]:
        d = match_date + timedelta(days=d_offset)
        if d not in lookup:
            continue
        matches = lookup[d]

        # Exact
        for m in matches:
            if m["home_norm"] == home_n and m["away_norm"] == away_n:
                return m["hg"], m["ag"]

        # Substring
        for m in matches:
            h_sub = (home_n in m["home_norm"] or m["home_norm"] in home_n) and len(min(home_n, m["home_norm"], key=len)) > 4
            a_sub = (away_n in m["away_norm"] or m["away_norm"] in away_n) and len(min(away_n, m["away_norm"], key=len)) > 4
            if h_sub and a_sub:
                return m["hg"], m["ag"]

        # Fuzzy
        best_score = 0
        best_match = None
        for m in matches:
            h_score = SequenceMatcher(None, home_n, m["home_norm"]).ratio()
            a_score = SequenceMatcher(None, away_n, m["away_norm"]).ratio()
            combined = (h_score + a_score) / 2
            if combined > best_score:
                best_score = combined
                best_match = m

        if best_score >= 0.55:
            return best_match["hg"], best_match["ag"]

    return None, None


# -------------------------------------------------------------
# Main
# -------------------------------------------------------------

def main():
    logger.info("Starting results updater...")

    if not MASTER_PATH.exists():
        logger.error(f"Master spreadsheet not found: {MASTER_PATH}")
        return

    # Load spreadsheet and find rows needing results
    wb = openpyxl.load_workbook(MASTER_PATH)
    ws = wb[MASTER_SHEET]

    rows_to_fill = []
    needed_dates = set()

    for r in range(2, ws.max_row + 1):
        bt = ws.cell(row=r, column=1).value
        if bt is None:
            break
        n_val = ws.cell(row=r, column=14).value
        o_val = ws.cell(row=r, column=15).value
        if n_val is not None and n_val != "" and o_val is not None and o_val != "":
            continue

        d = ws.cell(row=r, column=2).value
        comp = ws.cell(row=r, column=5).value
        home = ws.cell(row=r, column=3).value
        away = ws.cell(row=r, column=4).value
        pred = ws.cell(row=r, column=8).value

        if isinstance(d, datetime):
            d = d.date()
        if d is None or d >= date.today():
            continue

        rows_to_fill.append((r, bt, d, home, away, pred, comp))
        # Fetch ±1 day to cover Perth/Europe TZ skew
        needed_dates.add(d)
        needed_dates.add(d - timedelta(days=1))
        needed_dates.add(d + timedelta(days=1))

    logger.info(f"Rows needing results: {len(rows_to_fill)}")
    wb.close()

    if not rows_to_fill:
        logger.info("Nothing to update")
        return

    def _apply(lookup, rows):
        filled, remaining = 0, []
        for r, bt, d, home, away, pred, comp in rows:
            hg, ag = find_result(lookup, d, home, away)
            if hg is None:
                remaining.append((r, bt, d, home, away, pred, comp))
                continue
            ws.cell(row=r, column=14, value=hg)
            ws.cell(row=r, column=15, value=ag)
            filled += 1
        return filled, remaining

    # Re-open for writing
    wb = openpyxl.load_workbook(MASTER_PATH)
    ws = wb[MASTER_SHEET]

    filled_rows = 0
    remaining = rows_to_fill

    # --- Primary: Fotmob (global, single endpoint) ---
    logger.info("Primary source: Fotmob")
    try:
        fotmob_df = download_fotmob_results(needed_dates)
    except Exception as e:
        logger.warning(f"Fotmob fetch failed: {e}")
        fotmob_df = pd.DataFrame()

    if not fotmob_df.empty:
        lookup = build_lookup(fotmob_df)
        n, remaining = _apply(lookup, remaining)
        filled_rows += n
        logger.info(f"Fotmob filled: {n}, remaining: {len(remaining)}")

    # --- Fallback: football-data.co.uk + ESPN (only if anything still missing) ---
    if remaining:
        logger.info(f"Fallback: football-data.co.uk + ESPN for {len(remaining)} unresolved rows")
        fd_df = download_football_data_csvs()

        espn_needed = defaultdict(set)
        for _r, _bt, d, _h, _a, _p, comp in remaining:
            code = ESPN_LEAGUES.get(comp)
            if code:
                espn_needed[code].add(d)
                espn_needed[code].add(d - timedelta(days=1))
                espn_needed[code].add(d + timedelta(days=1))

        espn_df = download_espn_results(espn_needed) if espn_needed else pd.DataFrame()

        dfs = [df for df in [fd_df, espn_df] if not df.empty]
        if dfs:
            lookup = build_lookup(pd.concat(dfs, ignore_index=True))
            n, remaining = _apply(lookup, remaining)
            filled_rows += n
            logger.info(f"Fallback filled: {n}, still unresolved: {len(remaining)}")

    logger.info(f"Total filled rows (N home, O away): {filled_rows}")
    if remaining:
        logger.warning(f"Unresolved rows: {len(remaining)}")
        for r, bt, d, home, away, _p, comp in remaining[:20]:
            logger.warning(f"  row {r}: {d} {home} vs {away} [{comp}]")

    sort_master_rows(ws)
    rebuild_stake_formulas(ws)

    wb.save(MASTER_PATH)
    logger.info(f"Saved: {MASTER_PATH}")


# Input columns (everything else is a formula and should not be touched)
INPUT_COLS = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 14, 15]


def sort_master_rows(ws):
    """Sort data rows in-place by match date (col B) then match time (col F).

    Only swaps values in input columns; formula columns (L, M, P-U) are left
    alone so their same-row references recompute against the new inputs, and
    the cumulative-range formulas (T, U) naturally reflect the sorted order.
    """
    last_row = 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is None:
            break
        last_row = r

    if last_row < 3:
        return

    rows = []
    for r in range(2, last_row + 1):
        vals = {c: ws.cell(row=r, column=c).value for c in INPUT_COLS}
        rows.append(vals)

    def sort_key(v):
        d = v.get(2)
        if isinstance(d, datetime):
            d = d.date()
        d_key = d if isinstance(d, date) else date.max
        mt = v.get(6)
        if isinstance(mt, datetime):
            t_key = mt.time()
        elif hasattr(mt, "hour") and hasattr(mt, "minute"):
            t_key = mt
        else:
            from datetime import time as _time
            t_key = _time(0, 0)
        return (d_key, t_key)

    rows.sort(key=sort_key)

    for i, vals in enumerate(rows):
        r = i + 2
        for c in INPUT_COLS:
            ws.cell(row=r, column=c, value=vals[c])

    logger.info(f"Sorted {len(rows)} rows by match date/time")


def _rpd(o365, bf):
    try:
        a, b = float(o365), float(bf)
        if a > b:
            return 1.0
        pct = abs(a - b) / ((a + b) / 2) * 100
        return 1.0 if pct < 1 else round(pct, 3)
    except (TypeError, ValueError, ZeroDivisionError):
        return None


def _is_core(bt, pred, vol, bf, rpd):
    if bt not in ("1.5G", "3.5G", "BTTS"):
        return False
    if pred != 0:
        return False
    try:
        vol = float(vol); bf = float(bf)
    except (TypeError, ValueError):
        return False
    if not (40 <= vol <= 1100):
        return False
    if bf <= 1.45 or rpd is None:
        return False
    if bf <= 2.7 and rpd > 2.8:
        return False
    if bf > 2.7 and rpd > 3.5:
        return False
    return True


def rebuild_stake_formulas(ws):
    """Rewrite Q (Stake) for every data row with precomputed BTTS-conflict
    and double-stake flags derived from the CURRENT row data. Must be called
    after any row sort so flags stay aligned with the correct fixtures."""
    from bet_tracker_updater import stake_formula  # local import to avoid cycle

    last_row = 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is None:
            break
        last_row = r
    if last_row < 2:
        return

    rows = []
    for r in range(2, last_row + 1):
        rows.append({
            "r": r,
            "bt": ws.cell(row=r, column=1).value,
            "d": ws.cell(row=r, column=2).value,
            "home": ws.cell(row=r, column=3).value,
            "away": ws.cell(row=r, column=4).value,
            "pred": ws.cell(row=r, column=8).value,
            "o365": ws.cell(row=r, column=9).value,
            "bf": ws.cell(row=r, column=10).value,
            "vol": ws.cell(row=r, column=11).value,
        })
    for x in rows:
        x["rpd"] = _rpd(x["o365"], x["bf"])

    pred_lookup = {
        (x["d"], x["home"], x["away"], x["bt"]): x["pred"] for x in rows
    }
    conflict_rows = set()
    # No conflict filters active

    core_rows = set()
    match_count = defaultdict(int)
    for x in rows:
        if x["r"] in conflict_rows:
            continue
        if _is_core(x["bt"], x["pred"], x["vol"], x["bf"], x["rpd"]):
            core_rows.add(x["r"])
            match_count[(x["d"], x["home"], x["away"])] += 1

    # Double-stake: RPD=1.0 + ≥2 core bets on same match, only highest BF per match
    match_dbl_candidates = defaultdict(list)
    for x in rows:
        if x["r"] not in core_rows:
            continue
        if x["rpd"] == 1.0 and match_count[(x["d"], x["home"], x["away"])] >= 2:
            try:
                bf_f = float(x["bf"])
            except (TypeError, ValueError):
                bf_f = 0
            match_dbl_candidates[(x["d"], x["home"], x["away"])].append((x["r"], bf_f))

    double_rows = set()
    for mk, candidates in match_dbl_candidates.items():
        best_r = max(candidates, key=lambda c: c[1])[0]
        double_rows.add(best_r)

    for x in rows:
        r = x["r"]
        ws.cell(row=r, column=17, value=stake_formula(
            r,
            is_conflict=(r in conflict_rows),
            is_double_stake=(r in double_rows),
        ))

    logger.info(
        f"Rebuilt stake formulas: {len(core_rows)} core, "
        f"{len(double_rows)} double, {len(conflict_rows)} conflicts"
    )


if __name__ == "__main__":
    main()