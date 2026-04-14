"""
Backfill missing Home Goals (N) / Away Goals (O) in the master spreadsheet.

Strategy:
1. Read spreadsheet, find all rows where N or O is empty.
2. Group unique fixtures by (date, home, away, competition).
3. For each unique fixture, look up the result via:
   a) football-data.co.uk season CSVs (current + previous seasons)
   b) ESPN scoreboard API (per league + date)
4. Fuzzy-match team names, write hg/ag to ALL rows for that fixture.
5. Save spreadsheet, log unresolved fixtures to logs/backfill_unresolved.csv.

Reuses the league mappings + matching logic from results_updater.py.
"""

import csv
import logging
import sys
import time
from collections import defaultdict
from datetime import datetime, date, timedelta
from io import StringIO
from pathlib import Path

import openpyxl
import pandas as pd
import requests

PROJECT_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(PROJECT_ROOT))

from results_updater import (
    MASTER_PATH,
    MASTER_SHEET,
    MAIN_CSVS,
    NEW_CSVS,
    ESPN_LEAGUES,
    normalize,
    build_lookup,
    find_result,
)

LOG_DIR = PROJECT_ROOT / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)
LOG_PATH = LOG_DIR / f"backfill_goals_{datetime.now().strftime('%Y-%m-%d_%H%M')}.log"
UNRESOLVED_CSV = LOG_DIR / "backfill_unresolved.csv"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(str(LOG_PATH)), logging.StreamHandler()],
    force=True,
)
logger = logging.getLogger("backfill_goals")


# Past seasons to query for football-data.co.uk MAIN league CSVs
# (current 2526 is in MAIN_CSVS already)
PAST_SEASONS = ["2425", "2324", "2223"]


def fd_main_url(current_url, season):
    """Replace 2526 in a football-data MAIN URL with another season code."""
    return current_url.replace("/2526/", f"/{season}/")


def download_all_fd_csvs():
    """Download football-data.co.uk results across current + past seasons."""
    all_results = []

    # Current season (MAIN + NEW) — same logic as results_updater
    for league, url in MAIN_CSVS.items():
        for season_url in [url] + [fd_main_url(url, s) for s in PAST_SEASONS]:
            try:
                r = requests.get(season_url, timeout=20)
                if r.status_code != 200:
                    continue
                df = pd.read_csv(StringIO(r.text))
                if "FTHG" not in df.columns:
                    continue
                results = pd.DataFrame({
                    "_date": pd.to_datetime(df["Date"], dayfirst=True, errors="coerce").dt.date,
                    "_home": df["HomeTeam"],
                    "_away": df["AwayTeam"],
                    "_hg": pd.to_numeric(df["FTHG"], errors="coerce"),
                    "_ag": pd.to_numeric(df["FTAG"], errors="coerce"),
                }).dropna(subset=["_date", "_hg", "_ag"])
                all_results.append(results)
                logger.info(f"FD {league} ({season_url[-12:]}): {len(results)}")
            except Exception as e:
                logger.warning(f"FD {league} {season_url}: {e}")
            time.sleep(0.2)

    for league, url in NEW_CSVS.items():
        try:
            r = requests.get(url, timeout=20)
            r.raise_for_status()
            df = pd.read_csv(StringIO(r.text))
            results = pd.DataFrame({
                "_date": pd.to_datetime(df["Date"], dayfirst=True, errors="coerce").dt.date,
                "_home": df["Home"],
                "_away": df["Away"],
                "_hg": pd.to_numeric(df["HG"], errors="coerce"),
                "_ag": pd.to_numeric(df["AG"], errors="coerce"),
            }).dropna(subset=["_date", "_hg", "_ag"])
            all_results.append(results)
            logger.info(f"FD-new {league}: {len(results)}")
        except Exception as e:
            logger.warning(f"FD-new {league}: {e}")

    if not all_results:
        return pd.DataFrame()
    return pd.concat(all_results, ignore_index=True)


def download_espn_for(needed):
    """Fetch ESPN scoreboard for {league_code: set(dates)}.
    ESPN returns ~one day at a time for soccer scoreboards."""
    rows = []
    for code, dates in needed.items():
        for d in sorted(dates):
            url = (
                f"https://site.api.espn.com/apis/site/v2/sports/soccer/{code}"
                f"/scoreboard?dates={d.strftime('%Y%m%d')}"
            )
            try:
                r = requests.get(url, timeout=15)
                r.raise_for_status()
                data = r.json()
                for ev in data.get("events", []):
                    comps = ev.get("competitions", [])
                    if not comps:
                        continue
                    teams = comps[0].get("competitors", [])
                    if len(teams) != 2:
                        continue
                    home = away = None
                    for t in teams:
                        info = {
                            "name": t["team"]["displayName"],
                            "score": t.get("score", ""),
                        }
                        if t.get("homeAway") == "home":
                            home = info
                        else:
                            away = info
                    if home and away and home["score"] != "" and away["score"] != "":
                        try:
                            rows.append({
                                "_date": d,
                                "_home": home["name"],
                                "_away": away["name"],
                                "_hg": int(home["score"]),
                                "_ag": int(away["score"]),
                            })
                        except ValueError:
                            pass
            except Exception as e:
                logger.debug(f"ESPN {code} {d}: {e}")
            time.sleep(0.15)
    logger.info(f"ESPN: {len(rows)} results fetched")
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def main():
    logger.info(f"Loading {MASTER_PATH}")
    wb = openpyxl.load_workbook(MASTER_PATH)
    ws = wb[MASTER_SHEET]

    # Pass 1: find rows missing N/O
    fixtures = {}  # (date, home, away) -> {"comp": ..., "rows": [r, ...]}
    for r in range(2, ws.max_row + 1):
        bt = ws.cell(row=r, column=1).value
        if bt is None:
            break
        n_val = ws.cell(row=r, column=14).value
        o_val = ws.cell(row=r, column=15).value
        if n_val not in (None, "") and o_val not in (None, ""):
            continue

        d = ws.cell(row=r, column=2).value
        home = ws.cell(row=r, column=3).value
        away = ws.cell(row=r, column=4).value
        comp = ws.cell(row=r, column=5).value
        if isinstance(d, datetime):
            d = d.date()
        if not isinstance(d, date):
            continue
        # Skip future / today
        if d >= date.today():
            continue

        key = (d, home, away)
        if key not in fixtures:
            fixtures[key] = {"comp": comp, "rows": []}
        fixtures[key]["rows"].append(r)

    logger.info(f"Unique unresolved fixtures: {len(fixtures)}")
    affected_rows = sum(len(v["rows"]) for v in fixtures.values())
    logger.info(f"Affected rows: {affected_rows}")

    if not fixtures:
        logger.info("Nothing to backfill")
        return

    # Build ESPN-needed map
    espn_needed = defaultdict(set)
    for (d, _h, _a), meta in fixtures.items():
        code = ESPN_LEAGUES.get(meta["comp"])
        if code:
            # Query +/- 1 day to handle TZ shift
            espn_needed[code].add(d)
            espn_needed[code].add(d - timedelta(days=1))
            espn_needed[code].add(d + timedelta(days=1))

    logger.info(f"ESPN league/date pairs to fetch: "
                f"{sum(len(v) for v in espn_needed.values())} "
                f"across {len(espn_needed)} leagues")

    # Download
    logger.info("Downloading football-data.co.uk CSVs (current + past seasons)...")
    fd_df = download_all_fd_csvs()
    logger.info(f"FD total: {len(fd_df)}")

    logger.info("Downloading ESPN results...")
    espn_df = download_espn_for(espn_needed)

    dfs = [df for df in [fd_df, espn_df] if not df.empty]
    if not dfs:
        logger.error("No results downloaded — aborting")
        return
    all_results = pd.concat(dfs, ignore_index=True)
    logger.info(f"Total result records available: {len(all_results)}")

    lookup = build_lookup(all_results)

    # Pass 2: resolve and write
    resolved = 0
    rows_filled = 0
    unresolved = []

    for (d, home, away), meta in fixtures.items():
        hg, ag = find_result(lookup, d, home, away)
        if hg is None:
            unresolved.append({
                "date": d.isoformat(),
                "home": home,
                "away": away,
                "competition": meta["comp"],
                "rows": ";".join(str(r) for r in meta["rows"]),
            })
            continue
        for r in meta["rows"]:
            ws.cell(row=r, column=14, value=hg)
            ws.cell(row=r, column=15, value=ag)
            rows_filled += 1
        resolved += 1

    logger.info(f"Resolved fixtures: {resolved}/{len(fixtures)}")
    logger.info(f"Rows filled: {rows_filled}")
    logger.info(f"Unresolved fixtures: {len(unresolved)}")

    if unresolved:
        with open(UNRESOLVED_CSV, "w", newline="") as f:
            writer = csv.DictWriter(
                f, fieldnames=["date", "home", "away", "competition", "rows"]
            )
            writer.writeheader()
            writer.writerows(unresolved)
        logger.info(f"Unresolved list: {UNRESOLVED_CSV}")

    logger.info("Saving spreadsheet...")
    wb.save(MASTER_PATH)
    logger.info(f"Saved {MASTER_PATH}")


if __name__ == "__main__":
    main()